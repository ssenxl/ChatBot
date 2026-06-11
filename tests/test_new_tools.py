"""เทส tools ใหม่: suggest_week, compare_weeks, export_excel"""
import json
import os
from datetime import datetime, timedelta

import pytest

import data_cache as dc_mod
import response_processor as rp
from response_processor import ResponseProcessor, _yw_from_date


def _yw(k: int) -> str:
    return _yw_from_date(datetime.now() + timedelta(weeks=k))


@pytest.fixture
def proc_with_cache():
    """inject mc/kg_ava CSV ปลอมเข้า DataCache singleton แล้วคืน ResponseProcessor"""
    y1, y2, y3, y4 = _yw(1), _yw(2), _yw(3), _yw(4)
    mc_csv = "\n".join([
        "YW,Group,Guage,Total,Used_N,Used_F,Ava",
        f"{y1},SKP,24,10,5,0,5",      # ก่อน min plannable — ต้องถูกกรองออก
        f"{y2},SKP,24,10,10,0,0",     # min plannable แต่เต็ม
        f"{y2},SKP,28,8,8,0,0",
        f"{y3},SKP,24,10,8,0,2",
        f"{y3},SKPLE,24,6,5,0,1",
        f"{y4},SKP,28,8,5,0,3",
    ])
    kg_csv = "\n".join([
        "YW,Group,KG_Ava",
        f"{y3},SKP,1500.5",
        f"{y3},SKPLE,800",
        f"{y4},SKP,2000",
    ])
    dc = dc_mod.get_data_cache()
    with dc._lock:
        old_cache = dc._cache.get('query_cap_ava')
        old_ready = dc._ready.get('query_cap_ava', False)
        dc._cache['query_cap_ava'] = {'success': True, 'data': {'mc': mc_csv, 'booking': '', 'kg_ava': kg_csv}}
        dc._ready['query_cap_ava'] = True
    yield ResponseProcessor(), (y1, y2, y3, y4)
    with dc._lock:
        if old_cache is None:
            dc._cache.pop('query_cap_ava', None)
        else:
            dc._cache['query_cap_ava'] = old_cache
        dc._ready['query_cap_ava'] = old_ready


# ---------- suggest_week ----------
def test_suggest_week_finds_earliest_free_week(proc_with_cache):
    proc, (y1, y2, y3, y4) = proc_with_cache
    out = proc._tool_suggest_week(group="SKP")
    lines = out.splitlines()
    assert lines[0] == "YW,Ava,KG_Ava,Groups"
    # week แรกที่ว่างของ SKP คือ y3 (Ava=2) — y1 ถูกกรอง, y2 เต็ม
    assert lines[1].startswith(f"{y3},2,1500.50,SKP")
    assert f"สัปดาห์เร็วที่สุดที่ลงได้คือ YW {y3}" in out


def test_suggest_week_excludes_weeks_before_min_plannable(proc_with_cache):
    proc, (y1, *_rest) = proc_with_cache
    out = proc._tool_suggest_week(group="SKP")
    assert y1 not in out  # y1 มี Ava 5 แต่ก่อน current+2


def test_suggest_week_respects_machines_needed(proc_with_cache):
    proc, (_y1, _y2, y3, y4) = proc_with_cache
    out = proc._tool_suggest_week(group="SKP", machines_needed=3)
    lines = out.splitlines()
    assert lines[1].startswith(f"{y4},3")  # y3 มีแค่ 2 เครื่อง ไม่พอ


def test_suggest_week_gauge_filter_omits_kg(proc_with_cache):
    proc, (_y1, _y2, y3, _y4) = proc_with_cache
    out = proc._tool_suggest_week(group="SKP", gauge="24")
    lines = out.splitlines()
    assert lines[1] == f"{y3},2,,SKP"  # KG_Ava เว้นว่างเมื่อ filter gauge


def test_suggest_week_no_group_aggregates_all(proc_with_cache):
    proc, (_y1, _y2, y3, _y4) = proc_with_cache
    out = proc._tool_suggest_week()
    lines = out.splitlines()
    # y3: SKP(2) + SKPLE(1) = 3, KG = 1500.5 + 800
    assert lines[1] == f"{y3},3,2300.50,SKP; SKPLE"


def test_suggest_week_unknown_group(proc_with_cache):
    proc, _ = proc_with_cache
    out = proc._tool_suggest_week(group="NOPE")
    assert "ไม่พบสัปดาห์" in out


# ---------- compare_weeks ----------
def test_compare_weeks_delta_and_pct(proc_with_cache):
    proc, (_y1, _y2, y3, y4) = proc_with_cache
    out = proc._tool_compare_weeks(week_a=y3, week_b=y4, group="SKP")
    lines = out.splitlines()
    assert lines[0] == f"Group,Total_{y3},Ava_{y3},Total_{y4},Ava_{y4},Ava_Delta,Ava_Pct"
    assert lines[1] == f"SKP,10,2,8,3,+1,50.0%"
    assert "ห้ามคำนวณเอง" in out


def test_compare_weeks_zero_base_has_no_pct(proc_with_cache):
    proc, (_y1, y2, y3, _y4) = proc_with_cache
    out = proc._tool_compare_weeks(week_a=y2, week_b=y3, group="SKP")
    skp_line = [l for l in out.splitlines() if l.startswith("SKP,")][0]
    assert skp_line.endswith(",+2,")  # Ava_a = 0 → ไม่มี %


def test_compare_weeks_totals_row_for_multiple_groups(proc_with_cache):
    proc, (_y1, _y2, y3, y4) = proc_with_cache
    out = proc._tool_compare_weeks(week_a=y3, week_b=y4)
    total_lines = [l for l in out.splitlines() if l.startswith("(รวม),")]
    assert len(total_lines) == 1
    # y3: ava 2+1=3 | y4: ava 3 → delta +0
    assert ",3," in total_lines[0] and ",+0," in total_lines[0]


def test_compare_weeks_same_week_rejected(proc_with_cache):
    proc, (_y1, _y2, y3, _y4) = proc_with_cache
    out = proc._tool_compare_weeks(week_a=y3, week_b=y3)
    assert "สัปดาห์เดียวกัน" in out


def test_compare_weeks_unparseable_week(proc_with_cache):
    proc, _ = proc_with_cache
    out = proc._tool_compare_weeks(week_a="abc", week_b="xyz")
    assert "ไม่เข้าใจสัปดาห์" in out


def test_parse_week_arg_relative_and_numeric():
    assert ResponseProcessor._parse_week_arg("202624") == "202624"
    assert ResponseProcessor._parse_week_arg("อีก 3 สัปดาห์") == _yw(3)
    assert ResponseProcessor._parse_week_arg("ไม่ใช่สัปดาห์") is None
    assert ResponseProcessor._parse_week_arg(None) is None


# ---------- export_excel ----------
def _cleanup(fname):
    path = os.path.join(ResponseProcessor._EXCEL_EXPORT_DIR, fname)
    if os.path.exists(path):
        os.remove(path)


def test_export_excel_creates_file_and_link():
    proc = ResponseProcessor()
    out = proc._tool_export_excel(json.dumps({
        "title": "แผนทอ SKP",
        "filename": "knit_plan_skp",
        "columns": ["Item", "KP Weight (kg)", "สัปดาห์"],
        "rows": [["F100114/10A0", 1200.5, "202626"], ["FD4BASPI80B0", 800, "202627"]],
    }))
    assert "[📥 ดาวน์โหลด knit_plan_skp_" in out
    fname = out.split("(/download/excel/")[1].rstrip(")")
    path = os.path.join(ResponseProcessor._EXCEL_EXPORT_DIR, fname)
    try:
        assert os.path.exists(path)
        from openpyxl import load_workbook
        ws = load_workbook(path).active
        assert ws.cell(1, 1).value == "แผนทอ SKP"       # title แถวแรก
        assert ws.cell(3, 1).value == "Item"             # header แถว 3 (เว้น 1 แถว)
        assert ws.cell(4, 2).value == 1200.5             # ตัวเลขเก็บเป็น number
        assert ws.cell(4, 3).value == "202626"           # week code เก็บเป็น string
    finally:
        _cleanup(fname)


def test_export_excel_sanitizes_filename():
    proc = ResponseProcessor()
    out = proc._tool_export_excel(json.dumps({
        "filename": "../evil name!!",
        "columns": ["A"],
        "rows": [[1]],
    }))
    fname = out.split("(/download/excel/")[1].rstrip(")")
    try:
        assert fname.startswith("evilname_")  # อักขระอันตรายถูกตัดออกหมด
        assert "/" not in fname and "\\" not in fname and ".." not in fname
    finally:
        _cleanup(fname)


def test_export_excel_rejects_empty():
    proc = ResponseProcessor()
    assert "ไม่สำเร็จ" in proc._tool_export_excel(json.dumps({"columns": [], "rows": []}))
    assert "ไม่สำเร็จ" in proc._tool_export_excel("{not json")


# ---------- export_excel: source_tool mode (server-side fetch) ----------
@pytest.fixture
def item_cache():
    """inject query_item CSV ปลอม (รูปแบบเดียวกับ _aggregate_item_plan)"""
    y3 = _yw(3)
    item_csv = "\n".join([
        "Item,Group,KP_Weight,YW",
        f"F100114/10A0,SBP,546.25,{y3}",
        f"FD4BASPI80B0,SBP,144.34,{y3}",
        f"FD5GI28B0,RL,300,{y3}",
    ])
    dc = dc_mod.get_data_cache()
    with dc._lock:
        old = dc._cache.get('query_item'), dc._ready.get('query_item', False)
        dc._cache['query_item'] = {'success': True, 'data': item_csv}
        dc._ready['query_item'] = True
    yield ResponseProcessor(), y3
    with dc._lock:
        if old[0] is None:
            dc._cache.pop('query_item', None)
        else:
            dc._cache['query_item'] = old[0]
        dc._ready['query_item'] = old[1]


def test_export_excel_source_tool_fetches_server_side(item_cache):
    proc, y3 = item_cache
    out = proc._tool_export_excel(json.dumps({
        "source_tool": "get_item_plan",
        "week": y3,
        "filename": f"item_plan_{y3}",
        "title": f"Item ทั้งหมดสัปดาห์ {y3}",
    }))
    assert "[📥 ดาวน์โหลด" in out, out
    fname = out.split("(/download/excel/")[1].rstrip(")")
    path = os.path.join(ResponseProcessor._EXCEL_EXPORT_DIR, fname)
    try:
        from openpyxl import load_workbook
        ws = load_workbook(path).active
        assert ws.cell(3, 1).value == "Item"           # header จาก CSV ของ tool
        assert ws.cell(4, 1).value == "F100114/10A0"   # item code เป็น string
        assert ws.cell(4, 3).value == 546.25           # KP_Weight เป็น number
        assert ws.cell(4, 4).value == y3               # YW คงเป็น string
        assert ws.max_row == 6                          # title(1) + ว่าง(1) + header(1) + 3 rows
    finally:
        _cleanup(fname)


def test_export_excel_source_tool_passes_through_not_found(item_cache):
    proc, _y3 = item_cache
    out = proc._tool_export_excel(json.dumps({
        "source_tool": "get_item_plan",
        "item_code": "NOPE999",
    }))
    assert "ไม่สำเร็จ" in out and "NOPE999" in out


def test_export_excel_unknown_source_tool():
    proc = ResponseProcessor()
    out = proc._tool_export_excel(json.dumps({"source_tool": "render_chart"}))
    assert "ไม่รู้จัก source_tool" in out


def test_csv_to_table_skips_notes_and_converts_types():
    text = "Item,Group,KP_Weight,YW\nF1,SKP,12.5,202626\n[TOTAL_KP_WEIGHT=12.50 kg]\n[หมายเหตุ: ...]"
    cols, rows = ResponseProcessor._csv_to_table(text)
    assert cols == ["Item", "Group", "KP_Weight", "YW"]
    assert rows == [["F1", "SKP", 12.5, "202626"]]


def test_csv_to_table_rejects_non_table():
    assert ResponseProcessor._csv_to_table("ไม่พบ item NOPE ใน Item Plan") == ([], [])
    assert ResponseProcessor._csv_to_table("") == ([], [])


# ---------- analyze_plan_impact ----------
@pytest.fixture
def impact_cache():
    """inject query_item + query_cap_ava สำหรับวิเคราะห์ผลกระทบ"""
    y3, y4 = _yw(3), _yw(4)
    item_csv = "\n".join([
        "Item,Group,KP_Weight,YW",
        f"FTARGET01A0,SKP,500,{y3}",
        f"FOTHER01A0,SKP,300,{y3}",
        f"FOTHER02A0,SKP,200,{y3}",
        f"FLONE01A0,RL,100,{y3}",     # อยู่กลุ่ม RL คนเดียว
        f"FTARGET01A0,SKP,400,{y4}",
    ])
    mc_csv = "\n".join([
        "YW,Group,Guage,Total,Used_N,Used_F,Ava",
        f"{y3},SKP,24,10,8,0,2",
        f"{y3},RL,24,4,1,0,3",
        f"{y4},SKP,24,10,5,0,5",
    ])
    dc = dc_mod.get_data_cache()
    with dc._lock:
        old_item = dc._cache.get('query_item'), dc._ready.get('query_item', False)
        old_cap = dc._cache.get('query_cap_ava'), dc._ready.get('query_cap_ava', False)
        dc._cache['query_item'] = {'success': True, 'data': item_csv}
        dc._ready['query_item'] = True
        dc._cache['query_cap_ava'] = {'success': True, 'data': {'mc': mc_csv, 'booking': '', 'kg_ava': ''}}
        dc._ready['query_cap_ava'] = True
    yield ResponseProcessor(), (y3, y4)
    with dc._lock:
        for key, old in (('query_item', old_item), ('query_cap_ava', old_cap)):
            if old[0] is None:
                dc._cache.pop(key, None)
            else:
                dc._cache[key] = old[0]
            dc._ready[key] = old[1]


def test_impact_lists_affected_items_sorted(impact_cache):
    proc, (y3, _y4) = impact_cache
    out = proc._tool_analyze_plan_impact(item_code="FTARGET01A0", week=y3)
    assert f"ถ้าถอดแผน FTARGET01A0 YW {y3} กลุ่ม SKP (500 kg)" in out
    assert "2 รายการ" in out and "รวม 500.00 kg" in out
    # เรียงตาม KP มาก→น้อย
    assert out.index("FOTHER01A0 (300 kg)") < out.index("FOTHER02A0 (200 kg)")


def test_impact_estimates_freed_machines(impact_cache):
    proc, (y3, _y4) = impact_cache
    out = proc._tool_analyze_plan_impact(item_code="FTARGET01A0", week=y3)
    # group_kp=1000, used=8 → แผนนี้ใช้ ~ 500/(1000/8) = 4.0 เครื่อง, Ava 2 → ~6.0
    assert "Total=10, Used=8, Ava=2" in out
    assert "~4.0 เครื่อง" in out and "~6.0" in out


def test_impact_no_other_items(impact_cache):
    proc, (y3, _y4) = impact_cache
    out = proc._tool_analyze_plan_impact(item_code="FLONE01A0")
    assert "ไม่กระทบ item อื่น" in out


def test_impact_analyzes_all_weeks_without_filter(impact_cache):
    proc, (y3, y4) = impact_cache
    out = proc._tool_analyze_plan_impact(item_code="FTARGET01A0")
    assert f"YW {y3}" in out and f"YW {y4}" in out


def test_impact_includes_min_plannable_note(impact_cache):
    proc, _ = impact_cache
    out = proc._tool_analyze_plan_impact(item_code="FTARGET01A0")
    assert rp._min_plannable_yw() in out


def test_impact_item_not_found(impact_cache):
    proc, _ = impact_cache
    assert "ไม่พบ item" in proc._tool_analyze_plan_impact(item_code="NOPE999")
    assert "ต้องระบุ item_code" in proc._tool_analyze_plan_impact()


def test_impact_substring_match(impact_cache):
    proc, _ = impact_cache
    out = proc._tool_analyze_plan_impact(item_code="TARGET01")
    assert "FTARGET01A0" in out


# ---------- item_capability ----------
@pytest.fixture
def capability_cache():
    """inject query_item_capability + query_cap_ava + query_item"""
    y2, y3 = _yw(2), _yw(3)
    capability_csv = "\n".join([
        "Item,Group,SubGroup,Cap",
        "FTARGET01A0,SKP,SKP 24G,1200",
        "FTARGET01A0,SKPLE,SKPLE 24G,900",
        "FTARGET01A0,RL,RL 28G,700",
        "FOTHER01A0,SKP,SKP 24G,1000",
    ])
    mc_csv = "\n".join([
        "YW,Group,Guage,Total,Used_N,Used_F,Ava",
        f"{y2},SKP,24,10,10,0,0",     # SKP เต็มที่ y2
        f"{y3},SKP,24,10,8,0,2",      # SKP ว่างครั้งแรก y3
        f"{y2},SKPLE,24,6,4,0,2",     # SKPLE ว่างตั้งแต่ y2
        f"{y2},RL,28,4,4,0,0",        # RL เต็มตลอด
        f"{y3},RL,28,4,4,0,0",
    ])
    item_csv = "\n".join([
        "Item,Group,KP_Weight,YW",
        f"FTARGET01A0,SKP,500,{y3}",  # มีแผนอยู่ที่ SKP เท่านั้น
    ])
    dc = dc_mod.get_data_cache()
    keys = {
        'query_item_capability': {'success': True, 'data': capability_csv},
        'query_cap_ava': {'success': True, 'data': {'mc': mc_csv, 'booking': '', 'kg_ava': ''}},
        'query_item': {'success': True, 'data': item_csv},
    }
    with dc._lock:
        old = {k: (dc._cache.get(k), dc._ready.get(k, False)) for k in keys}
        for k, v in keys.items():
            dc._cache[k] = v
            dc._ready[k] = True
    yield ResponseProcessor(), (y2, y3)
    with dc._lock:
        for k, (oc, orr) in old.items():
            if oc is None:
                dc._cache.pop(k, None)
            else:
                dc._cache[k] = oc
            dc._ready[k] = orr


def test_capability_lists_groups_with_next_free(capability_cache):
    proc, (y2, y3) = capability_cache
    out = proc._tool_item_capability(item_code="FTARGET01A0")
    lines = out.splitlines()
    assert lines[0] == "Group,SubGroup,Cap,Planned,Next_Free_YW,Next_Free_Ava"
    assert f"SKP,SKP 24G,1200,yes,{y3},2" in out      # มีแผนอยู่ + ว่างครั้งแรก y3
    assert f"SKPLE,SKPLE 24G,900,,{y2},2" in out      # ไม่มีแผน + ว่างตั้งแต่ y2
    assert "RL,RL 28G,700,,ไม่มีว่าง,0" in out          # เต็มตลอด
    assert "ทอได้ 3 กลุ่ม" in out


def test_capability_specific_week_shows_ava(capability_cache):
    proc, (y2, _y3) = capability_cache
    out = proc._tool_item_capability(item_code="FTARGET01A0", week=y2)
    assert f"Ava_{y2}" in out.splitlines()[0]
    assert "SKP,SKP 24G,1200,yes,0" in out
    assert "SKPLE,SKPLE 24G,900,,2" in out


def test_capability_substring_match(capability_cache):
    proc, _ = capability_cache
    out = proc._tool_item_capability(item_code="TARGET01")
    assert "FTARGET01A0" in out and "ทอได้ 3 กลุ่ม" in out


def test_capability_not_found(capability_cache):
    proc, _ = capability_cache
    assert "ไม่พบ item" in proc._tool_item_capability(item_code="NOPE999")
    assert "ต้องระบุ item_code" in proc._tool_item_capability()


def test_aggregate_item_capability_dedups():
    rows = [
        {'Item': 'F1', 'Group': 'SKP', 'SubGroup': 'SKP 24G', 'Cap': 1200, 'MAP2': 'x'},
        {'Item': 'F1', 'Group': 'SKP', 'SubGroup': 'SKP 24G', 'Cap': 1200, 'MAP2': 'y'},  # ซ้ำ
        {'Item': 'F1', 'Group': 'RL', 'SubGroup': '', 'Cap': None},
        {'Item': '', 'Group': 'SKP'},   # ไม่มี item — ข้าม
    ]
    csv_text = dc_mod._aggregate_item_capability(rows)
    lines = csv_text.splitlines()
    assert lines[0] == "Item,Group,SubGroup,Cap"
    assert len(lines) == 3
    assert "F1,SKP,SKP 24G,1200" in lines
    assert "F1,RL,," in lines


# ---------- get_sales_summary ----------
@pytest.fixture
def sales_cache():
    y3, y4 = _yw(3), _yw(4)
    sales_data = {
        'somchai': {'name': 'Somchai', 'item_count': 10, 'kg': 5000.0,
                    'kg_yw': {y3: 3000.0, y4: 2000.0}},
        'gift': {'name': 'Gift', 'item_count': 4, 'kg': 8000.0,
                 'kg_yw': {y3: 8000.0}},
        '__system__': {'name': 'ระบบทั้งหมด', 'item_count': 14, 'kg': 13000.0},
    }
    dc = dc_mod.get_data_cache()
    with dc._lock:
        old = dc._cache.get('query_sales'), dc._ready.get('query_sales', False)
        dc._cache['query_sales'] = {'success': True, 'data': sales_data}
        dc._ready['query_sales'] = True
    yield ResponseProcessor(), (y3, y4)
    with dc._lock:
        if old[0] is None:
            dc._cache.pop('query_sales', None)
        else:
            dc._cache['query_sales'] = old[0]
        dc._ready['query_sales'] = old[1]


def test_sales_ranking_all(sales_cache):
    proc, _ = sales_cache
    out = proc._tool_get_sales_summary()
    lines = out.splitlines()
    assert lines[0] == "Sales,Items_All,KG"
    assert lines[1] == "Gift,4,8000.0"        # เรียงตาม kg มาก→น้อย
    assert lines[2] == "Somchai,10,5000.0"
    assert "(รวม 2 sales),,13000.0" in out
    assert "__system__" not in out


def test_sales_by_name_partial_match(sales_cache):
    proc, _ = sales_cache
    out = proc._tool_get_sales_summary(sales_name="gif")
    assert "Gift,4,8000.0" in out


def test_sales_week_filter_uses_kg_yw(sales_cache):
    proc, (y3, _y4) = sales_cache
    out = proc._tool_get_sales_summary(sales_name="somchai", week=y3)
    assert "Somchai,10,3000.0" in out
    assert f"YW {y3}" in out


def test_sales_not_found_lists_names(sales_cache):
    proc, _ = sales_cache
    out = proc._tool_get_sales_summary(sales_name="nobody")
    assert "ไม่พบ sales" in out and "Gift" in out


def test_sales_old_cache_without_kg_yw_falls_back(sales_cache):
    proc, (y3, _y4) = sales_cache
    dc = dc_mod.get_data_cache()
    with dc._lock:
        dc._cache['query_sales'] = {'success': True, 'data': {
            'somchai': {'name': 'Somchai', 'item_count': 10, 'kg': 5000.0},
        }}
    out = proc._tool_get_sales_summary(sales_name="somchai", week=y3)
    assert "Somchai,10,5000.0" in out
    assert "ยังไม่พร้อม แสดงยอดรวม" in out


# ---------- group_utilization ----------
def test_utilization_ranking_tightest_first(proc_with_cache):
    proc, (_y1, y2, y3, _y4) = proc_with_cache
    # y2: SKP total 18 used 18 ava 0 (100%) | y3: SKP 10/8, SKPLE 6/5
    out = proc._tool_group_utilization(week=y2)
    lines = out.splitlines()
    assert lines[0] == "Group,Total,Used,Ava,Used_Pct,KG_Ava,Worst_YW,Worst_Ava"
    assert lines[1].startswith(f"SKP,18,18,0,100.0%")
    assert "ห้ามคำนวณเอง" in out


def test_utilization_sort_by_available(proc_with_cache):
    proc, (_y1, _y2, y3, _y4) = proc_with_cache
    out = proc._tool_group_utilization(week=y3, sort_by="available")
    lines = out.splitlines()
    # y3: SKP ava 2 > SKPLE ava 1
    assert lines[1].startswith("SKP,")
    assert lines[2].startswith("SKPLE,")


def test_utilization_worst_week_tracked(proc_with_cache):
    proc, (_y1, y2, y3, y4) = proc_with_cache
    # ครอบคลุม y1..y4 ด้วย weeks_ahead → worst ของ SKP = y2 (ava 0)
    out = proc._tool_group_utilization(weeks_ahead=5)
    skp_line = [l for l in out.splitlines() if l.startswith("SKP,")]
    assert skp_line and f",{y2}," in skp_line[0]


def test_utilization_no_data_in_range(proc_with_cache):
    proc, _ = proc_with_cache
    out = proc._tool_group_utilization(week="209901")
    assert "ไม่พบข้อมูล capacity" in out


# ---------- export nudge (กัน LLM reject คำขอ export) ----------
def test_export_nudge_triggers_on_excel_request():
    for msg in ["ขอเป็นไฟล์ excel ได้ไหม", "export ให้หน่อย", "ขอไฟล์หน่อยค่ะ", "ดาวน์โหลดข้อมูลนี้", "ขอเป็น xlsx"]:
        nudge = rp._export_nudge(msg)
        assert len(nudge) == 1, msg
        assert nudge[0]["role"] == "system"
        assert "export_excel" in nudge[0]["content"]


def test_export_nudge_silent_on_normal_questions():
    for msg in ["ข้อมูลเครื่องจักร", "capacity SKP สัปดาห์นี้", "สวัสดีค่ะ", "item F100114 อยู่กลุ่มไหน"]:
        assert rp._export_nudge(msg) == [], msg
    assert rp._export_nudge("") == []
    assert rp._export_nudge(None) == []


# ---------- dispatch ----------
def test_execute_tool_call_dispatches_new_tools(proc_with_cache, monkeypatch):
    import types
    proc, _ = proc_with_cache
    called = {}
    monkeypatch.setattr(proc, "_tool_suggest_week", lambda **kw: (called.setdefault("suggest", kw), "ok")[1])
    monkeypatch.setattr(proc, "_tool_compare_weeks", lambda **kw: (called.setdefault("compare", kw), "ok")[1])
    monkeypatch.setattr(proc, "_tool_export_excel", lambda raw: (called.setdefault("excel", raw), "ok")[1])

    def _call(name, args):
        fake = types.SimpleNamespace(function=types.SimpleNamespace(name=name, arguments=json.dumps(args)))
        return proc._execute_tool_call(fake)

    assert _call("suggest_week", {"group": "SKP", "machines_needed": 2}) == "ok"
    assert called["suggest"] == {"group": "SKP", "gauge": None, "machines_needed": 2}
    assert _call("compare_weeks", {"week_a": "202624", "week_b": "202626"}) == "ok"
    assert called["compare"]["week_a"] == "202624"
    assert _call("export_excel", {"columns": ["A"], "rows": [[1]]}) == "ok"
    assert json.loads(called["excel"]) == {"columns": ["A"], "rows": [[1]]}
