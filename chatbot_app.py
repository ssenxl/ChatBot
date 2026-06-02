from collections import defaultdict
from functools import wraps
import asyncio
import os
from pathlib import Path
import threading
import time

from datetime import datetime as _dt, timezone as _tz, timedelta as _td

from flask import Flask, flash, jsonify, redirect, render_template, request, session, url_for
from flask.json.provider import DefaultJSONProvider
from flask_wtf.csrf import CSRFProtect
from openpyxl import load_workbook

from database import Database
from mcp_client import get_mcp_client
from intent_detector import get_intent_detector
from response_processor import get_response_processor
from suggestion_engine import get_suggestion_engine
from data_cache import get_data_cache
from morning_greeting import MorningGreetingScheduler

_BKK = _tz(_td(hours=7))


class _BkkJSON(DefaultJSONProvider):
    def default(self, o):
        if isinstance(o, _dt) and o.tzinfo is None:
            return o.replace(tzinfo=_BKK).isoformat()
        return super().default(o)


app = Flask(__name__)
app.json_provider_class = _BkkJSON
app.json = _BkkJSON(app)
app.config['TEMPLATES_AUTO_RELOAD'] = True
csrf = CSRFProtect(app)

_secret_key = os.environ.get('FLASK_SECRET_KEY')
if not _secret_key:
    import warnings
    warnings.warn(
        "FLASK_SECRET_KEY is not set. Using a random key — all sessions will be lost on restart. "
        "Set FLASK_SECRET_KEY in your environment for production.",
        stacklevel=1,
    )
    _secret_key = os.urandom(32)
app.secret_key = _secret_key

db = Database()
mcp_client = get_mcp_client()
intent_detector = get_intent_detector()
response_processor = get_response_processor()
suggestion_engine = get_suggestion_engine()
morning_scheduler = MorningGreetingScheduler(db)

DEFAULT_CONVERSATION_TITLE = 'แชทใหม่'

_rate_lock = threading.Lock()
_rate_attempts: dict = defaultdict(list)
_RATE_WINDOW = 300   # 5 นาที
_RATE_MAX_LOGIN = 10
_RATE_MAX_FORGOT = 5


def _is_rate_limited(ip: str, bucket: str = 'login') -> bool:
    max_attempts = _RATE_MAX_FORGOT if bucket == 'forgot' else _RATE_MAX_LOGIN
    now = time.time()
    key = f"{bucket}:{ip}"
    with _rate_lock:
        _rate_attempts[key] = [t for t in _rate_attempts[key] if now - t < _RATE_WINDOW]
        if len(_rate_attempts[key]) >= max_attempts:
            return True
        _rate_attempts[key].append(now)
        return False
REQUIRED_SESSION_KEYS = ('user_id', 'username', 'user_role', 'user_email')
SIGNUP_EMAIL_COLUMN = os.getenv('SIGNUP_EMAIL_COLUMN', 'EMPLOYEE_EMAIL').strip() or 'EMPLOYEE_EMAIL'
SIGNUP_EMAIL_SOURCE_FILE = Path(
    os.getenv('SIGNUP_EMAIL_SOURCE_FILE', Path(__file__).resolve().parent / 'User_login' / 'Sales Email Update.xlsx')
)


def _parse_allowed_signup_emails():
    raw_value = os.getenv('ALLOWED_SIGNUP_EMAILS', '')
    if not raw_value.strip():
        return set()
    return {
        email.strip().lower()
        for email in raw_value.split(',')
        if email.strip()
    }


def _load_allowed_signup_emails_from_excel(file_path, email_column):
    if not file_path.exists():
        print(f"Warning: Signup email source file not found: {file_path}")
        return set()

    allowed_emails = set()
    workbook = load_workbook(file_path, data_only=True, read_only=True)

    try:
        for worksheet in workbook.worksheets:
            header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                continue

            headers = [str(value).strip() if value is not None else '' for value in header_row]
            if email_column not in headers:
                continue

            email_index = headers.index(email_column)

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                if email_index >= len(row):
                    continue

                raw_email = row[email_index]
                if raw_email is None:
                    continue

                normalized_email = str(raw_email).strip().lower()
                if normalized_email:
                    allowed_emails.add(normalized_email)
    finally:
        workbook.close()

    return allowed_emails


ALLOWED_SIGNUP_EMAILS = _load_allowed_signup_emails_from_excel(SIGNUP_EMAIL_SOURCE_FILE, SIGNUP_EMAIL_COLUMN)
ALLOWED_SIGNUP_EMAILS.update(_parse_allowed_signup_emails())

# MCP initialization flag
_mcp_initialized = False

def initialize_mcp():
    """Initialize MCP servers"""
    global _mcp_initialized
    if not _mcp_initialized:
        try:
            asyncio.run(mcp_client.initialize_servers())
            _mcp_initialized = True
            print("MCP servers initialized successfully")
        except Exception as e:
            print(f"Warning: Failed to initialize MCP servers: {e}")

# Initialize MCP and start data cache on app startup
with app.app_context():
    try:
        initialize_mcp()
    except Exception as e:
        print(f"MCP initialization skipped: {e}")
    try:
        get_data_cache().start()
        print("DataCache started — pre-loading booking data in background...")
    except Exception as e:
        print(f"DataCache start failed: {e}")
    try:
        morning_scheduler.start()
        print("MorningGreetingScheduler started — daily greeting at 08:00")
    except Exception as e:
        print(f"MorningGreetingScheduler start failed: {e}")
    # ตรวจสอบ default password ที่ยังไม่ถูกเปลี่ยน
    import warnings as _w
    from werkzeug.security import check_password_hash as _cph
    try:
        for _default in [('admin', 'adminscm'), ('user', 'user123')]:
            _u = db.get_user_by_username(_default[0])
            if _u:
                _row = db.get_connection()
                _c = _row.cursor()
                _c.execute("SELECT password_hash FROM users WHERE username = %s", (_default[0],))
                _h = _c.fetchone()
                _row.close()
                if _h and _cph(_h['password_hash'], _default[1]):
                    _w.warn(
                        f"SECURITY: ผู้ใช้ '{_default[0]}' ยังใช้รหัสผ่าน default '{_default[1]}' — กรุณาเปลี่ยนทันที!",
                        stacklevel=1
                    )
    except Exception:
        pass


def has_valid_session():
    return all(session.get(key) for key in REQUIRED_SESSION_KEYS)


@app.after_request
def set_security_headers(response):
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.tailwindcss.com https://cdnjs.cloudflare.com; "
        "style-src 'self' 'unsafe-inline' https://cdnjs.cloudflare.com https://fonts.googleapis.com; "
        "font-src 'self' https://cdnjs.cloudflare.com https://fonts.gstatic.com; "
        "img-src 'self' data:;"
    )
    return response


@app.before_request
def track_user_activity():
    """อัปเดต last_activity ทุก request และตรวจ session_version — ถ้า version ไม่ตรงจะ kick ออกทันที"""
    if has_valid_session():
        user_id = session.get('user_id')
        if user_id:
            try:
                valid = db.update_last_activity(user_id, session.get('session_version', 1))
                if not valid:
                    session.clear()
                    return redirect(url_for('login'))
            except Exception:
                pass  # ไม่ให้ error ใน tracking กระทบ request หลัก


def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not has_valid_session():
            session.clear()
            return redirect(url_for('login'))
        if session.get('user_role') != 'admin':
            return jsonify({'success': False, 'message': 'ไม่มีสิทธิ์เข้าถึง'}), 403
        return f(*args, **kwargs)
    return decorated_function


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not has_valid_session():
            session.clear()
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def generate_conversation_title(message):
    cleaned = ' '.join((message or '').strip().split())
    if not cleaned:
        return DEFAULT_CONVERSATION_TITLE
    return cleaned[:60] + ('...' if len(cleaned) > 60 else '')


def generate_ai_response(message):
    message_lower = (message or '').strip().lower()

    thai_greetings = ['สวัสดี', 'หวัดดี']
    eng_greetings = ['hello', 'hi']
    if any(keyword in message_lower for keyword in thai_greetings + eng_greetings):
        is_english = any(keyword in message_lower for keyword in eng_greetings) and not any(keyword in message_lower for keyword in thai_greetings)
        if is_english:
            greeting_msg = "Hello! I'm I-SAVE Chatbot. Feel free to ask me anything about the I-SAVE system. I'm happy to help!"
        else:
            greeting_msg = 'สวัสดีค่ะ น้อง I-SAVE Chatbot ค่ะพี่ๆ สามารถสอบถามข้อมูล หรือพิมพ์คำถามที่ต้องการได้เลยนะคะ น้องยินดีช่วยเหลือค่ะ'
        return {
            'message': greeting_msg,
            'type': 'text'
        }

    if any(keyword in message_lower for keyword in ['ตั้งหัวข้อ', 'เปลี่ยนหัวข้อ', 'rename title', 'rename']):
        return {
            'message': (
                'ถ้าต้องการเปลี่ยนชื่อหัวข้อ ให้กดปุ่มแก้ชื่อหัวข้อด้านบนของแชทได้เลยครับ '
                'หัวข้อจะถูกเก็บไว้ตามผู้ใช้ และกลับมาเปิดอ่านใหม่ได้ในภายหลัง'
            ),
            'type': 'text'
        }

    if any(keyword in message_lower for keyword in ['แชทใหม่', 'new chat', 'เริ่มใหม่', 'เปลี่ยนเรื่อง']):
        return {
            'message': 'ถ้าต้องการเปลี่ยนเรื่องสนทนา ให้กดปุ่ม "แชทใหม่" ในแถบด้านซ้าย แล้วเริ่มหัวข้อใหม่ได้ทันทีครับ',
            'type': 'text'
        }

    return {
        'message': (
            f'รับทราบครับ\n\n'
            f'ประเด็นที่คุณพิมพ์คือ: "{message}"\n\n'
            'หากต้องการให้ช่วยต่อแบบมีโครงสร้าง ลองระบุเพิ่มได้เช่น:\n'
            '1. เป้าหมายที่ต้องการ\n'
            '2. ข้อมูลหรือบริบทที่เกี่ยวข้อง\n'
            '3. รูปแบบคำตอบที่ต้องการ เช่น สรุป, แผนงาน, ตัวอย่างข้อความ, หรือรายการขั้นตอน'
        ),
        'type': 'text'
    }


@app.route('/')
def index():
    if has_valid_session():
        return redirect(url_for('chat'))
    session.clear()
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if _is_rate_limited(request.remote_addr, 'login'):
            flash('คุณพยายาม login บ่อยเกินไป กรุณารอ 5 นาทีแล้วลองใหม่', 'error')
            return render_template('login.html')

        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash('กรุณากรอกชื่อผู้ใช้และรหัสผ่าน', 'error')
            return render_template('login.html')

        user = db.authenticate_user(username, password)

        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['user_role'] = user['role']
            session['user_email'] = user['email']
            session['session_version'] = user['session_version']

            db.update_last_login(user['id'])
            db.log_activity(user['id'], 'login', {
                'username': username,
                'ip_address': request.remote_addr,
                'user_agent': request.headers.get('User-Agent')
            })

            flash(f'ยินดีต้อนรับ {username}!', 'success')
            return redirect(url_for('chat'))

        flash('ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง', 'error')

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    allowed_signup_email_count = len(ALLOWED_SIGNUP_EMAILS)

    def render_register_page():
        return render_template(
            'register.html',
            allowed_signup_email_count=allowed_signup_email_count
        )

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not all([username, email, password, confirm_password]):
            flash('กรุณากรอกข้อมูลให้ครบถ้วน', 'error')
            return render_register_page()

        if ALLOWED_SIGNUP_EMAILS and email.lower() not in ALLOWED_SIGNUP_EMAILS:
            flash('อีเมลนี้ไม่ได้รับอนุญาตให้สมัครสมาชิก', 'error')
            return render_register_page()

        if password != confirm_password:
            flash('รหัสผ่านไม่ตรงกัน', 'error')
            return render_register_page()

        user_id = db.create_user(username, email, password)

        if user_id:
            db.log_activity(user_id, 'register', {
                'username': username,
                'email': email,
                'ip_address': request.remote_addr
            })
            flash('สมัครสมาชิกสำเร็จ! กรุณาล็อกอิน', 'success')
            return redirect(url_for('login'))

        flash('ชื่อผู้ใช้หรืออีเมลนี้มีอยู่แล้ว', 'error')

    return render_register_page()


@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        if _is_rate_limited(request.remote_addr, 'forgot'):
            flash('คุณพยายามรีเซ็ตรหัสผ่านบ่อยเกินไป กรุณารอสักครู่แล้วลองใหม่', 'error')
            return render_template('forgot_password.html')

        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip().lower()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        if not username or not email:
            flash('กรุณากรอกชื่อผู้ใช้และอีเมล', 'error')
            return render_template('forgot_password.html')

        if not new_password or not confirm_password:
            flash('กรุณากรอกรหัสผ่านใหม่และยืนยันรหัสผ่าน', 'error')
            return render_template('forgot_password.html')

        if new_password != confirm_password:
            flash('รหัสผ่านใหม่ไม่ตรงกัน', 'error')
            return render_template('forgot_password.html')

        if len(new_password) < 6:
            flash('รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร', 'error')
            return render_template('forgot_password.html')

        user = db.get_user_by_username(username)
        # ตรวจสอบว่า username และ email ตรงกับบัญชีเดียวกัน
        if user and user.get('email', '').lower() == email:
            db.update_user_password(user['id'], new_password)
            db.log_activity(user['id'], 'password_reset', {
                'ip_address': request.remote_addr,
                'reset_method': 'forgot_password'
            })
            flash('รีเซ็ตรหัสผ่านสำเร็จ! กรุณาเข้าสู่ระบบด้วยรหัสผ่านใหม่', 'success')
            return redirect(url_for('login'))

        # คำตอบเหมือนกันทั้งกรณีหา user ไม่เจอและ email ไม่ตรง (ป้องกัน user enumeration)
        flash('ชื่อผู้ใช้หรืออีเมลไม่ถูกต้อง', 'error')

    return render_template('forgot_password.html')


@app.route('/logout')
def logout():
    if 'user_id' in session:
        db.log_activity(session['user_id'], 'logout', {
            'username': session.get('username'),
            'ip_address': request.remote_addr
        })

    session.clear()
    flash('ออกจากระบบสำเร็จ', 'info')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    return redirect(url_for('chat'))


@app.route('/chat')
@login_required
def chat():
    _try_send_morning_greeting(session['user_id'])
    return render_template('chat_enhanced.html')


def _try_send_morning_greeting(user_id: int):
    """ส่ง morning greeting เมื่อ user เปิด chat หลัง 08:00 ถ้ายังไม่ได้รับวันนี้"""
    try:
        now = _dt.now(_BKK)
        if now.hour < 8:
            return
        if db.has_morning_greeting_today(user_id):
            return
        user = db.get_user_by_username(session.get('username', ''))
        if not user or user.get('role') == 'admin':
            return

        from data_cache import get_data_cache as _gdc
        _cache = _gdc()
        mc_cached, mc_ready = _cache.get('query_machine')
        mc_data    = mc_cached.get('data', {}) if mc_ready and mc_cached else {}
        mc_csv     = mc_data.get('mc', '')
        kg_ava_csv = mc_data.get('kg_ava', '')
        booking_cached, _ = _cache.get('query_booking')
        booking_csv = booking_cached.get('data', '') if booking_cached else ''
        item_cached, _   = _cache.get('query_item')
        item_csv   = item_cached.get('data', '') if item_cached else ''

        from morning_greeting import build_greeting_text as _bgt
        text = _bgt(session.get('username', ''), mc_csv, kg_ava_csv, booking_csv, item_csv)
        today = now.date()
        conv_id = db.create_conversation(user_id, f"I-SAVE News {today.day:02d}/{today.month:02d}")
        db.add_message(conv_id, 'assistant', text, 'text')
        db.record_morning_greeting(user_id, conv_id)
    except Exception as e:
        import logging as _log
        _log.getLogger(__name__).warning(f"_try_send_morning_greeting failed: {e}")


@app.route('/user-info')
@login_required
def user_info():
    return jsonify({
        'success': True,
        'user': {
            'id': session['user_id'],
            'username': session['username'],
            'email': session['user_email'],
            'role': session['user_role']
        }
    })


@app.route('/conversations')
@login_required
def get_conversations():
    conversations = db.get_user_conversations(session['user_id'])
    return jsonify({'success': True, 'conversations': conversations})


@app.route('/conversations', methods=['POST'])
@login_required
def create_conversation():
    data = request.get_json(silent=True) or {}
    title = (data.get('title') or '').strip() or DEFAULT_CONVERSATION_TITLE
    user_id = session['user_id']

    conversation_id = db.create_conversation(user_id, title)
    db.log_activity(user_id, 'create_conversation', {
        'conversation_id': conversation_id,
        'title': title
    })

    return jsonify({
        'success': True,
        'conversation_id': conversation_id,
        'title': title
    })


@app.route('/conversations/<int:conversation_id>/messages')
@login_required
def get_messages(conversation_id):
    user_id = session['user_id']
    conversation = db.get_conversation(conversation_id, user_id)
    if not conversation:
        return jsonify({'success': False, 'message': 'ไม่พบหัวข้อสนทนา'}), 404

    messages = db.get_conversation_messages(conversation_id, user_id)
    return jsonify({
        'success': True,
        'conversation': conversation,
        'messages': messages
    })


@app.route('/conversations/<int:conversation_id>/messages', methods=['POST'])
@login_required
def send_message(conversation_id):
    user_id = session['user_id']
    username = session.get('username', 'คุณ')
    conversation = db.get_conversation(conversation_id, user_id)
    if not conversation:
        return jsonify({'success': False, 'message': 'ไม่พบหัวข้อสนทนา'}), 404

    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    message_type = data.get('message_type', 'text')

    if not message:
        return jsonify({'success': False, 'message': 'กรุณากรอกข้อความ'}), 400

    # ดึง history ก่อนบันทึกข้อความปัจจุบัน เพื่อให้ได้ context ที่ถูกต้อง
    conv_history = db.get_conversation_messages(conversation_id, user_id)

    # บันทึกข้อความของผู้ใช้
    user_message_id = db.add_message(conversation_id, 'user', message, message_type)

    if conversation['title'] == DEFAULT_CONVERSATION_TITLE:
        db.update_conversation_title(conversation_id, user_id, generate_conversation_title(message))

    # ประมวลผลด้วยระบบใหม่ (Intent Detection + MCP + Suggestions)
    start_time = time.time()

    try:
        # Process message พร้อม conversation history เพื่อรองรับ multi-turn context
        processed_response = asyncio.run(
            response_processor.process_message(message, username, conv_history)
        )
    except Exception as e:
        import traceback
        print(f"[ERROR] process_message failed: {e}")
        traceback.print_exc()
        processed_response = None

    if processed_response is None:
        ai_response = generate_ai_response(message)
        ai_message_id = db.add_message(
            conversation_id, 'assistant',
            ai_response['message'], ai_response['type']
        )
        ai_response['message_id'] = ai_message_id
    else:
        response_metadata = {
            'intent': processed_response.metadata.get('intent'),
            'confidence': processed_response.metadata.get('confidence'),
            'processing_path': processed_response.processing_path,
            'mcp_calls_count': len(processed_response.mcp_calls) if processed_response.mcp_calls else 0
        }

        # บันทึก AI response ก่อน (สำคัญที่สุด)
        ai_message_id = db.add_message(
            conversation_id, 'assistant',
            processed_response.message,
            processed_response.response_type,
            response_metadata
        )

        # Logging แยก try/except ไม่ให้พังทั้ง request
        try:
            intent_log_id = db.log_intent(
                conversation_id=conversation_id,
                message_id=user_message_id,
                user_message=message,
                intent=processed_response.metadata.get('intent', 'unknown'),
                confidence=processed_response.metadata.get('confidence', 0.0),
                matched_keywords=processed_response.metadata.get('matched_keywords', []),
                processing_path=processed_response.processing_path
            )
        except Exception as e:
            print(f"[WARN] log_intent failed: {e}")
            intent_log_id = None

        try:
            if processed_response.mcp_calls:
                for mcp_call in processed_response.mcp_calls:
                    tool_parts = mcp_call['tool'].split('/')
                    mcp_server = tool_parts[0]
                    tool_name_log = tool_parts[1] if len(tool_parts) > 1 else mcp_call['tool']
                    db.log_mcp_interaction(
                        conversation_id=conversation_id,
                        message_id=user_message_id,
                        intent_log_id=intent_log_id,
                        mcp_server=mcp_server,
                        tool_name=tool_name_log,
                        tool_arguments={},
                        tool_result={'success': True},  # ไม่เก็บ raw data ขนาดใหญ่
                        success=True,
                        execution_time_ms=int((time.time() - start_time) * 1000)
                    )
        except Exception as e:
            print(f"[WARN] log_mcp_interaction failed: {e}")

        try:
            if processed_response.processing_path == 'shortcut' and processed_response.suggestions:
                suggestions = processed_response.suggestions
            else:
                suggestions = suggestion_engine.generate_suggestions(
                    current_intent=processed_response.metadata.get('intent', 'unknown'),
                    conversation_history=db.get_conversation_messages(conversation_id, user_id),
                    user_context={'role': session.get('user_role', 'user')}
                )
            db.save_suggestions(conversation_id, ai_message_id, suggestions)
        except Exception as e:
            print(f"[WARN] suggestions failed: {e}")
            suggestions = []

        try:
            meta = processed_response.metadata or {}
            if meta.get('total_tokens'):
                db.log_token_usage(
                    user_id=user_id,
                    conversation_id=conversation_id,
                    message_id=ai_message_id,
                    model=meta.get('model', 'unknown'),
                    prompt_tokens=meta.get('prompt_tokens', 0),
                    completion_tokens=meta.get('completion_tokens', 0),
                    total_tokens=meta.get('total_tokens', 0),
                    tool_calls_count=len(processed_response.mcp_calls) if processed_response.mcp_calls else 0,
                    response_time_ms=int((time.time() - start_time) * 1000),
                )
        except Exception as e:
            print(f"[WARN] log_token_usage failed: {e}")

        try:
            db.log_activity(user_id, 'send_message', {
                'conversation_id': conversation_id,
                'message_type': message_type,
                'message_length': len(message),
                'intent': processed_response.metadata.get('intent'),
                'processing_path': processed_response.processing_path,
                'response_time_ms': int((time.time() - start_time) * 1000)
            })
        except Exception as e:
            print(f"[WARN] log_activity failed: {e}")

        ai_response = {
            'message': processed_response.message,
            'type': processed_response.response_type,
            'metadata': response_metadata,
            'suggestions': suggestions,
            'data': processed_response.data,
            'message_id': ai_message_id
        }

    updated_conversation = db.get_conversation(conversation_id, user_id)
    return jsonify({
        'success': True,
        'conversation': updated_conversation,
        'ai_response': ai_response
    })


@app.route('/conversations/<int:conversation_id>', methods=['PUT'])
@login_required
def update_conversation(conversation_id):
    user_id = session['user_id']
    data = request.get_json(silent=True) or {}
    title = (data.get('title') or '').strip()

    if not title:
        return jsonify({'success': False, 'message': 'กรุณาระบุชื่อหัวข้อ'}), 400

    success = db.update_conversation_title(conversation_id, user_id, title)
    if not success:
        return jsonify({'success': False, 'message': 'ไม่พบหัวข้อสนทนา'}), 404

    db.log_activity(user_id, 'update_conversation_title', {
        'conversation_id': conversation_id,
        'new_title': title
    })
    return jsonify({'success': True, 'message': 'อัปเดตหัวข้อสำเร็จ'})


@app.route('/conversations/<int:conversation_id>', methods=['DELETE'])
@login_required
def delete_conversation(conversation_id):
    user_id = session['user_id']
    success = db.delete_conversation(conversation_id, user_id)

    if success:
        db.log_activity(user_id, 'delete_conversation', {
            'conversation_id': conversation_id
        })
        return jsonify({'success': True})

    return jsonify({'success': False, 'message': 'ไม่พบหัวข้อสนทนา'}), 404


@app.route('/change-password', methods=['POST'])
@login_required
def change_password():
    data = request.get_json(silent=True) or {}
    old_password = (data.get('old_password') or '').strip()
    new_password = (data.get('new_password') or '').strip()

    if not old_password or not new_password:
        return jsonify({'success': False, 'message': 'กรุณากรอกข้อมูลให้ครบถ้วน'}), 400
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'รหัสผ่านใหม่ต้องมีอย่างน้อย 6 ตัวอักษร'}), 400

    success = db.change_password(session['user_id'], old_password, new_password)
    if not success:
        return jsonify({'success': False, 'message': 'รหัสผ่านเดิมไม่ถูกต้อง'}), 400

    db.log_activity(session['user_id'], 'change_password', {})
    return jsonify({'success': True, 'message': 'เปลี่ยนรหัสผ่านสำเร็จ'})


@app.route('/admin')
@admin_required
def admin_panel():
    return render_template('admin.html')


@app.route('/admin/users')
@admin_required
def admin_get_users():
    users = db.get_all_users()
    return jsonify({'success': True, 'users': users})


@app.route('/admin/users/<int:user_id>/conversations')
@admin_required
def admin_get_user_conversations(user_id):
    conversations = db.get_user_conversations(user_id, limit=100)
    return jsonify({'success': True, 'conversations': conversations})


@app.route('/admin/conversations/<int:conversation_id>/messages')
@admin_required
def admin_get_conversation_messages(conversation_id):
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute(
        'SELECT user_id FROM conversations WHERE id = %s AND is_active = TRUE',
        (conversation_id,)
    )
    row = cursor.fetchone()
    conn.close()
    if not row:
        return jsonify({'success': False, 'message': 'ไม่พบการสนทนา'}), 404
    messages = db.get_conversation_messages(conversation_id, row['user_id'])
    return jsonify({'success': True, 'messages': messages})


@app.route('/admin/users/<int:user_id>/toggle-active', methods=['POST'])
@admin_required
def admin_toggle_user_active(user_id):
    if user_id == session['user_id']:
        return jsonify({'success': False, 'message': 'ไม่สามารถปิดใช้งานตัวเองได้'}), 400
    data = request.get_json(silent=True) or {}
    is_active = bool(data.get('is_active', True))
    success = db.toggle_user_active(user_id, is_active)
    if not success:
        return jsonify({'success': False, 'message': 'ไม่พบผู้ใช้หรือไม่สามารถแก้ไข admin ได้'}), 404
    db.log_activity(session['user_id'], 'toggle_user_active', {'target_user_id': user_id, 'is_active': is_active})
    return jsonify({'success': True})


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def admin_reset_user_password(user_id):
    data = request.get_json(silent=True) or {}
    new_password = (data.get('new_password') or '').strip()

    if not new_password:
        return jsonify({'success': False, 'message': 'กรุณาระบุรหัสผ่านใหม่'}), 400
    if len(new_password) < 6:
        return jsonify({'success': False, 'message': 'รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร'}), 400

    success = db.admin_reset_password(user_id, new_password)
    if not success:
        return jsonify({'success': False, 'message': 'ไม่พบผู้ใช้'}), 404

    db.log_activity(session['user_id'], 'admin_reset_password', {'target_user_id': user_id})
    return jsonify({'success': True, 'message': 'รีเซ็ตรหัสผ่านสำเร็จ'})


@app.route('/admin/token-usage')
@admin_required
def admin_token_usage():
    rows = db.get_token_usage_per_user()
    return jsonify({'success': True, 'rows': rows})


@app.route('/admin/feedback')
@admin_required
def admin_feedback():
    data = db.get_feedback_summary()
    return jsonify({'success': True, **data})


@app.route('/admin/cache/status')
@admin_required
def admin_cache_status():
    return jsonify({'success': True, **get_data_cache().get_status()})


@app.route('/admin/cache/refresh', methods=['POST'])
@admin_required
def admin_cache_refresh():
    """บังคับ refresh data cache จาก Power BI ทันที"""
    from data_cache import get_data_cache
    get_data_cache().force_refresh()
    return jsonify({'success': True, 'message': 'รีเฟรช cache เสร็จสิ้น'})


@app.route('/admin/morning-greeting/debug-sales')
@admin_required
def admin_debug_sales():
    """ตรวจสอบว่า query_sales cache มีข้อมูลอะไรบ้าง"""
    cached, ready = get_data_cache().get('query_sales')
    if not ready or not cached:
        return jsonify({'ready': False, 'message': 'query_sales ยังไม่พร้อม กรุณากด Refresh Cache ก่อน'})
    sales_map = cached.get('data', {})
    names = [v['name'] for k, v in sales_map.items() if k != '__system__']
    system = sales_map.get('__system__', {})
    return jsonify({
        'ready': True,
        'sales_count': len(names),
        'sales_names': names[:20],
        'system_total': system,
        'message': 'ถ้า sales_count = 0 แสดงว่าคอลัมน์ KNIT_SALE_NAME ไม่มีในข้อมูล หรือ cache ยังไม่ refresh'
    })


@app.route('/admin/morning-greeting/send', methods=['POST'])
@admin_required
def admin_send_morning_greeting():
    """ส่ง morning greeting ให้ทุก user ทันที (สำหรับทดสอบ)"""
    result = morning_scheduler.send_now()
    if result['sent'] == 0 and result['errors'] == 0:
        msg = 'ไม่มี user ที่ต้องส่ง (อาจส่งไปแล้ววันนี้ หรือยังไม่มี user ในระบบ)'
    else:
        names = ', '.join(result['usernames']) or '-'
        msg = f"ส่งสำเร็จ {result['sent']} คน ({names})"
        if result['errors']:
            msg += f", ผิดพลาด {result['errors']} คน"
    return jsonify({'success': True, 'message': msg, **result})


# New API Endpoints for Intent Detection & MCP

@app.route('/api/mcp/tools')
@login_required
def get_mcp_tools():
    """ดึงรายการ MCP tools ที่มีทั้งหมด"""
    tools = mcp_client.get_all_tools()
    tools_list = [
        {
            'name': tool.name,
            'description': tool.description,
            'server': tool.server_name,
            'input_schema': tool.input_schema
        }
        for tool in tools
    ]
    return jsonify({'success': True, 'tools': tools_list})


@app.route('/api/intents')
@login_required
def get_available_intents():
    """ดึงรายการ intents ที่มี"""
    intents = intent_detector.get_available_intents()
    return jsonify({'success': True, 'intents': intents})


@app.route('/api/quick-actions')
@login_required
def get_quick_actions():
    """ดึง quick actions สำหรับแสดงเป็นปุ่ม"""
    actions = suggestion_engine.get_quick_actions()
    return jsonify({'success': True, 'actions': actions})


@app.route('/conversations/<int:conversation_id>/analytics')
@login_required
def get_conversation_analytics(conversation_id):
    """ดึง analytics ของการสนทนา"""
    user_id = session['user_id']
    conversation = db.get_conversation(conversation_id, user_id)
    if not conversation:
        return jsonify({'success': False, 'message': 'ไม่พบหัวข้อสนทนา'}), 404
    
    # Intent logs
    intent_logs = db.get_intent_logs(conversation_id, limit=100)
    
    # MCP interactions
    mcp_logs = db.get_mcp_interactions(conversation_id, limit=100)
    
    # Suggestion analytics
    suggestion_analytics = db.get_suggestion_analytics(conversation_id=conversation_id)
    
    # Calculate statistics
    total_messages = len(intent_logs)
    avg_confidence = sum(log['confidence'] for log in intent_logs) / total_messages if total_messages > 0 else 0
    
    processing_paths = {}
    for log in intent_logs:
        path = log['processing_path']
        processing_paths[path] = processing_paths.get(path, 0) + 1
    
    intent_distribution = {}
    for log in intent_logs:
        intent = log['detected_intent']
        intent_distribution[intent] = intent_distribution.get(intent, 0) + 1
    
    return jsonify({
        'success': True,
        'analytics': {
            'total_messages': total_messages,
            'avg_confidence': round(avg_confidence, 2),
            'processing_paths': processing_paths,
            'intent_distribution': intent_distribution,
            'mcp_calls': len(mcp_logs),
            'suggestion_analytics': suggestion_analytics
        }
    })


@app.route('/messages/<int:message_id>/feedback', methods=['POST'])
@login_required
def message_feedback(message_id):
    data = request.get_json(silent=True) or {}
    feedback_type = (data.get('feedback_type') or '').strip()
    if feedback_type not in ('like', 'dislike'):
        return jsonify({'success': False, 'message': 'feedback_type ต้องเป็น like หรือ dislike'}), 400
    if not db.message_belongs_to_user(message_id, session['user_id']):
        return jsonify({'success': False, 'message': 'ไม่พบข้อความ'}), 404
    db.save_message_feedback(message_id, session['user_id'], feedback_type)
    return jsonify({'success': True})


@app.route('/api/suggestions/<int:suggestion_id>/click', methods=['POST'])
@login_required
def mark_suggestion_clicked(suggestion_id):
    if not db.suggestion_belongs_to_user(suggestion_id, session['user_id']):
        return jsonify({'success': False, 'message': 'ไม่พบ suggestion'}), 404
    try:
        db.mark_suggestion_clicked(suggestion_id)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


# Support Ticket routes
@app.route('/support/tickets', methods=['GET'])
@login_required
def get_my_tickets():
    tickets = db.get_user_tickets(session['user_id'])
    for t in tickets:
        t['created_at'] = str(t['created_at'])
        t['updated_at'] = str(t['updated_at'])
        t['user_read_at'] = str(t['user_read_at']) if t['user_read_at'] else None
    return jsonify({'success': True, 'tickets': tickets})


@app.route('/support/tickets', methods=['POST'])
@login_required
def create_ticket():
    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    subject = (data.get('subject') or '').strip()[:100]
    if not message:
        return jsonify({'success': False, 'message': 'กรุณาระบุข้อความ'}), 400
    ticket_id = db.create_support_ticket(session['user_id'], message, subject)
    db.log_activity(session['user_id'], 'create_support_ticket', {'ticket_id': ticket_id})
    return jsonify({'success': True, 'ticket_id': ticket_id})


@app.route('/support/tickets/<int:ticket_id>/replies', methods=['GET'])
@login_required
def get_ticket_replies(ticket_id):
    replies = db.get_ticket_replies(ticket_id, user_id=session['user_id'])
    if replies is None:
        return jsonify({'success': False, 'message': 'ไม่พบ ticket'}), 404
    return jsonify({'success': True, 'replies': replies})


@app.route('/support/tickets/<int:ticket_id>/replies', methods=['POST'])
@login_required
def add_user_reply(ticket_id):
    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    if not message:
        return jsonify({'success': False, 'message': 'กรุณาระบุข้อความ'}), 400
    ok = db.add_ticket_reply(ticket_id, 'user', message, user_id=session['user_id'])
    if not ok:
        return jsonify({'success': False, 'message': 'ไม่พบ ticket'}), 404
    return jsonify({'success': True})


@app.route('/support/unread-count')
@login_required
def support_unread_count():
    count = db.get_unread_reply_count(session['user_id'])
    return jsonify({'success': True, 'count': count})


@app.route('/admin/support/tickets')
@admin_required
def admin_get_tickets():
    tickets = db.get_all_tickets_admin()
    for t in tickets:
        t['created_at'] = str(t['created_at'])
        t['updated_at'] = str(t['updated_at'])
    return jsonify({'success': True, 'tickets': tickets})


@app.route('/admin/support/tickets/<int:ticket_id>/replies', methods=['GET'])
@admin_required
def admin_get_ticket_replies(ticket_id):
    replies = db.get_ticket_replies(ticket_id)
    if replies is None:
        return jsonify({'success': False, 'message': 'ไม่พบ ticket'}), 404
    return jsonify({'success': True, 'replies': replies})


@app.route('/admin/support/tickets/<int:ticket_id>/reply', methods=['POST'])
@admin_required
def admin_reply_ticket(ticket_id):
    data = request.get_json(silent=True) or {}
    message = (data.get('message') or '').strip()
    if not message:
        return jsonify({'success': False, 'message': 'กรุณาระบุข้อความ'}), 400
    db.add_ticket_reply(ticket_id, 'admin', message)
    db.log_activity(session['user_id'], 'admin_reply_ticket', {'ticket_id': ticket_id})
    return jsonify({'success': True})


@app.route('/admin/support/tickets/<int:ticket_id>/close', methods=['POST'])
@admin_required
def admin_close_ticket(ticket_id):
    db.close_ticket(ticket_id)
    return jsonify({'success': True})


if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'
    app.run(debug=debug_mode, host='0.0.0.0', port=5000, use_reloader=False)
