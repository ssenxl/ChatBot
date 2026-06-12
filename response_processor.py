import asyncio
import json
import os
import re
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional
from dataclasses import dataclass
import logging
from openai import OpenAI, AsyncOpenAI
from data_cache import get_data_cache

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MAX_TOOL_ITERATIONS = 5
OPENAI_MAX_RETRIES = 2  # จำนวนครั้งที่ลองเรียก OpenAI ซ้ำเมื่อ API พลาด (รวมครั้งแรกเป็น 3)
OPENAI_RETRY_BACKOFF = 0.8  # วินาที, หน่วงก่อน retry (เพิ่มขึ้นแบบ exponential)


@dataclass
class ProcessedResponse:
    message: str
    response_type: str  # 'text', 'data', 'chart', 'table'
    data: Optional[Dict] = None
    metadata: Optional[Dict] = None
    suggestions: Optional[List[str]] = None
    processing_path: str = 'agent'
    mcp_calls: Optional[List[Dict]] = None


def _clean_response(text: str) -> str:
    text = re.sub(r'\n{3,}', '\n\n', text)
    # Remove blank line between consecutive bullets
    _bullet_gap = re.compile(r'([ \t]*[-•*][^\n]+)\n\n([ \t]*[-•*])', re.MULTILINE)
    while _bullet_gap.search(text):
        text = _bullet_gap.sub(r'\1\n\2', text)
    # Remove blank line between non-bullet text and a bullet
    text = re.sub(r'([^\n]+)\n\n([ \t]*[-•*])', r'\1\n\2', text)
    # Remove blank line between a bullet and following non-bullet text
    text = re.sub(r'([ \t]*[-•*][^\n]+)\n\n([^-•*\n])', r'\1\n\2', text)
    return text.strip()


def _has_thai(text: str) -> bool:
    return bool(re.search(r'[฀-๿]', text))


def _has_english_words(text: str) -> bool:
    # Strip item-code-style tokens (must contain a digit or slash, e.g. F100114/10A0, SKP28G)
    # Pure uppercase words like "OK", "YES" are NOT stripped so they're detected as English.
    stripped = re.sub(r'\b[A-Z][A-Z0-9/\-]*(?=[0-9/])[A-Z0-9/\-]*\b', '', text)
    return bool(re.search(r'\b[a-zA-Z]{2,}\b', stripped))


def _detect_reply_language(message: str, history: Optional[List[Dict]] = None) -> str:
    """Return 'thai' or 'english'.
    - Any Thai character in message → Thai
    - No Thai + has English words → English
    - No signal → check recent history, default Thai
    """
    if _has_thai(message):
        return 'thai'
    if _has_english_words(message):
        return 'english'

    # No signal (e.g. bare item code) — fall back to recent history
    if history:
        for msg in reversed(history[-6:]):
            content = msg.get('content') or msg.get('message', '')
            if _has_thai(content):
                return 'thai'
            if _has_english_words(content):
                return 'english'

    return 'thai'


def _min_plannable_yw() -> str:
    future = datetime.now() + timedelta(weeks=2)
    iso = future.isocalendar()
    return f"{iso[0]}{iso[1]:02d}"


def _yw_from_date(d: datetime) -> str:
    iso = d.isocalendar()
    return f"{iso[0]}{iso[1]:02d}"


def _add_months(d: datetime, months: int) -> datetime:
    """เลื่อนเดือน รองรับข้ามปี (วันที่ 1 ของเดือนเป้าหมาย)"""
    m = d.month - 1 + months
    year = d.year + m // 12
    month = m % 12 + 1
    return datetime(year, month, 1)


def _yws_in_month(year: int, month: int) -> list:
    """คืน YW ทุกสัปดาห์ที่คาบเกี่ยวกับเดือนนั้น"""
    yws = []
    d = datetime(year, month, 1)
    while d.month == month:
        yw = _yw_from_date(d)
        if yw not in yws:
            yws.append(yw)
        d += timedelta(days=1)
    return yws


def _week_keyword_to_yw(text: str) -> list:
    """แปลงคำพูดเรื่องสัปดาห์/เดือนเป็น YW (YYYYWW) — รองรับทั้งแบบระบุเลขและแบบสัมพัทธ์
    เช่น 'สัปดาห์นี้/หน้า/ที่แล้ว', 'อีก 3 สัปดาห์', 'เดือนนี้/หน้า/ที่แล้ว',
    'this/next/last week', 'in 3 weeks', 'this/next/last month'."""
    now = datetime.now()
    year = now.year
    t = text.lower()
    yw_list = []

    def _add(yw):
        if yw not in yw_list:
            yw_list.append(yw)

    # --- relative weeks (อีก N / N ที่แล้ว) ---
    # จับ N-pattern ก่อน แล้วตัดส่วนที่จับได้ออกจาก t_single เพื่อไม่ให้ pattern เดี่ยวแมตช์ซ้อน
    t_single = t
    n_patterns = [
        (r'(?:อีก|in|after)\s*(\d{1,2})\s*(?:สัปดาห์|อาทิตย์|weeks?)', 1),
        (r'(\d{1,2})\s*(?:สัปดาห์|อาทิตย์|weeks?)\s*(?:ข้างหน้า|ถัดไป|จากนี้|from now|ahead|later)', 1),
        (r'(\d{1,2})\s*(?:สัปดาห์|อาทิตย์|weeks?)\s*(?:ที่แล้ว|ก่อน|ago|earlier)', -1),
    ]
    for pat, sign in n_patterns:
        for m in re.finditer(pat, t):
            _add(_yw_from_date(now + timedelta(weeks=sign * int(m.group(1)))))
        t_single = re.sub(pat, ' ', t_single)

    # --- single relative week (หาใน t_single ที่ตัด N-pattern ออกแล้ว) ---
    if re.search(r'สัปดาห์หน้า|อาทิตย์หน้า|สัปดาห์ถัดไป|อาทิตย์ถัดไป|next\s*week', t_single):
        _add(_yw_from_date(now + timedelta(weeks=1)))
    if re.search(r'สัปดาห์ที่แล้ว|อาทิตย์ที่แล้ว|สัปดาห์ก่อน|อาทิตย์ก่อน|last\s*week|previous\s*week', t_single):
        _add(_yw_from_date(now - timedelta(weeks=1)))
    if re.search(r'สัปดาห์นี้|อาทิตย์นี้|this\s*week', t_single):
        _add(_yw_from_date(now))

    # --- months (range of weeks) ---
    if re.search(r'เดือนนี้|this\s*month', t):
        for yw in _yws_in_month(now.year, now.month):
            _add(yw)
    if re.search(r'เดือนหน้า|เดือนถัดไป|next\s*month', t):
        nm = _add_months(now, 1)
        for yw in _yws_in_month(nm.year, nm.month):
            _add(yw)
    if re.search(r'เดือนที่แล้ว|เดือนก่อน|last\s*month|previous\s*month', t):
        pm = _add_months(now, -1)
        for yw in _yws_in_month(pm.year, pm.month):
            _add(yw)

    # --- explicit 6-digit YW (e.g. 202627) ---
    for m in re.finditer(r'\b(20\d{2})(\d{2})\b', text):
        wk = int(m.group(2))
        if 1 <= wk <= 53:
            _add(f"{m.group(1)}{wk:02d}")

    # --- explicit week number (week 22 / wk22 / w22 / สัปดาห์ที่ 22) ---
    patterns = [
        r'\bweek\s*(\d{1,2})\b',
        r'\bwk(\d{1,2})\b',
        r'\bw(\d{2})\b',              # "w22" (2-digit only) — avoids "w5", "w8" false positives
        r'สัปดาห์(?:ที่)?\s*(\d{1,2})',
    ]
    for pattern in patterns:
        for m in re.finditer(pattern, t):
            week_num = int(m.group(1))
            if 1 <= week_num <= 53:
                _add(f"{year}{week_num:02d}")

    return yw_list


def _group_matches(group_col: str, keywords: list) -> bool:
    for kw in keywords:
        if group_col == kw:
            return True
        if group_col.startswith(kw):
            rest = group_col[len(kw):]
            if rest and not rest[0].isalpha():
                return True
    return False


_EXPORT_REQUEST_RE = re.compile(
    r'excel|xlsx|export|ขอไฟล์|เป็นไฟล์|โหลดไฟล์|ขอเป็น\s*file|ดาวน์โหลดข้อมูล|download',
    re.IGNORECASE,
)


def _export_nudge(user_message: str) -> list:
    """system message เสริมเมื่อ user ขอ export ไฟล์ — LLM ตีความ rule 9 เองไม่เสถียร
    (บางครั้ง reject ว่าไม่ใช่หัวข้อ I-SAVE) จึงตรวจ keyword ฝั่ง Python แล้วสั่งตรงๆ"""
    if user_message and _EXPORT_REQUEST_RE.search(user_message):
        return [{"role": "system", "content": (
            "The user's current message is an EXPORT/FILE request for the I-SAVE data discussed in this conversation. "
            "This IS a supported I-SAVE topic — you MUST NOT reply with the rejection template from rule 9. "
            "Follow rule 16: call export_excel ONCE with source_tool + the SAME filters as the data just discussed "
            "(same group/week/item/gauge). Do NOT pass rows yourself. "
            "Include the returned download link line in your reply exactly as given."
        )}]
    return []


_TOOL_STATUS_LABELS = {
    'get_item_plan': 'กำลังดึงข้อมูล Item Plan...',
    'get_machine_capacity': 'กำลังดึงข้อมูล Capacity...',
    'get_booking': 'กำลังดึงข้อมูล Booking...',
    'get_knit_plan': 'กำลังดึงข้อมูลแผนทอ...',
    'suggest_week': 'กำลังหาสัปดาห์ที่มีเครื่องว่าง...',
    'compare_weeks': 'กำลังเปรียบเทียบสัปดาห์...',
    'analyze_plan_impact': 'กำลังวิเคราะห์ผลกระทบแผน...',
    'get_sales_summary': 'กำลังดึงยอด Sales...',
    'group_utilization': 'กำลังจัดอันดับการใช้เครื่อง...',
    'item_capability': 'กำลังดูกลุ่มที่ item ทอได้...',
    'export_excel': 'กำลังสร้างไฟล์ Excel...',
    'render_chart': 'กำลังสร้างกราฟ...',
    'render_table': 'กำลังสร้างตาราง...',
}

SYSTEM_PROMPT_TEMPLATE = """You are an AI assistant for the I-SAVE system at a textile factory.
คุณเป็นผู้ช่วย AI ของระบบ I-SAVE สำหรับโรงงานทอผ้า ชื่อว่า "น้อง I-SAVE Chatbot"

LANGUAGE RULE (highest priority — overrides ALL template responses below):
- Item codes (e.g. F100114/10A0), machine group names (e.g. SKP 28G), and technical codes are NOT language indicators — ignore them when detecting language.
- If the user's non-code text is Thai (or no non-code text exists) → reply in Thai
- If the user's non-code text is clearly English → reply in English
- If mixed → follow the dominant non-code language; default to Thai when unclear
- Always translate any template response to match the detected language.

คุณมี tools สำหรับดึงข้อมูลจากระบบ ให้เรียก tool ก่อนตอบทุกครั้งที่คำถามเกี่ยวกับข้อมูล:
- get_item_plan        : แผน item (กลุ่มเครื่อง, KP_Weight, สัปดาห์ที่วางแผน)
- get_machine_capacity : กำลังการผลิต (Total, Used_N, Used_F, Ava=available machines)
- get_booking          : การจองเครื่องต่อกลุ่มต่อสัปดาห์
- get_knit_plan        : แผนการทอ (item, กลุ่ม, KP_Weight ตามสัปดาห์)
- suggest_week         : หาสัปดาห์เร็วที่สุดที่มีเครื่องว่าง (กรองกฎ current+2 ให้อัตโนมัติ) — ใช้เมื่อถาม "ลงได้เร็วสุด week ไหน / ควรวางแผน week ไหน / มีที่ว่างเมื่อไหร่"
- compare_weeks        : เปรียบเทียบ capacity ระหว่าง 2 สัปดาห์ — คืน Delta และ % คำนวณให้แล้ว ห้ามคำนวณเอง
- analyze_plan_impact  : วิเคราะห์ผลกระทบถ้าถอด/เลื่อน/ย้ายแผนของ item — ใช้เมื่อถาม "ถ้าถอดแผนนี้กระทบ item อื่นไหม", "item นี้ใช้เครื่องร่วมกับใคร" (คืนค่าประมาณ + รายชื่อ item ที่กระทบให้แล้ว)
- get_sales_summary    : ยอดจอง (kg, จำนวน item) ต่อ sales — ใช้เมื่อถามยอดของ sales หรือจัดอันดับ sales (กรองช่วงเวลาได้)
- group_utilization    : จัดอันดับกลุ่มเครื่องตึง/ว่าง (Used% คำนวณให้แล้ว ห้ามคำนวณเอง) — ใช้เมื่อถาม "กลุ่มไหนแน่นสุด/ว่างสุด"
- item_capability      : กลุ่มที่ item "ทอได้" จาก Table_Item (ต่างจากกลุ่มที่ "มีแผน") + เครื่องว่างเร็วสุดของแต่ละกลุ่ม — ใช้เมื่อถาม "item นี้ทอได้ที่ไหนบ้าง / ย้ายกลุ่มได้ไหม"
- export_excel         : สร้างไฟล์ Excel ให้ user ดาวน์โหลด (เรียกเมื่อ user ขอ export/ไฟล์ Excel เท่านั้น)
- render_chart         : วาดกราฟ/แผนภูมิ (เรียกเมื่อ user ขอกราฟเท่านั้น)

กฎสำคัญ:
1. ตอบเฉพาะคำถามที่ user ถามใน message ปัจจุบันเท่านั้น — ห้ามนำคำถามหรือคำตอบจาก message ก่อนหน้ามาตอบซ้ำในรอบนี้ไม่ว่ากรณีใดทั้งสิ้น:
   - ห้ามเด็ดขาด: รวม/สรุป/ย้ำคำตอบเก่าไว้ในคำตอบปัจจุบัน แม้จะ "เพื่อความสะดวก" หรือ "สรุปให้ครบ"
   - ห้ามเด็ดขาด: ขึ้นต้นคำตอบด้วย "ขอตอบแยก 2 คำถาม" หรือ format ที่รวมหลายรอบคำถามเข้าด้วยกัน
   - ถ้า user ถามซ้ำหรือวนกลับมาถามคำถามเดิมใน message ใหม่ → ให้ call tool ใหม่เสมอ ห้ามนำคำตอบเก่าจาก history มาตอบแทน
   - ห้าม re-fetch ข้อมูลชุดเดิมในรอบเดียวกัน (same tool-call loop) เท่านั้น ไม่ใช่ข้ามรอบสนทนา
   - ถ้า message ใหม่ไม่ได้ถามเรื่องข้อมูล ห้ามเรียก tool เด็ดขาด
2. ถ้าคำถามเกี่ยวกับข้อมูลในระบบ ให้เรียก tool ก่อนเสมอ อย่าตอบจากความรู้ตัวเอง
3. ต้องการข้อมูลหลายอย่าง → เรียก tool ได้หลายครั้ง
4. YW = week code in format YYYYWW e.g. 202622 = year 2026 week 22
   - วันนี้คือ {today} (สัปดาห์ปัจจุบัน = YW {current_yw}). ใช้ค่านี้อ้างอิงเมื่อ user พูดถึงเวลาแบบสัมพัทธ์
   - เมื่อ user พูดถึงสัปดาห์/เดือนแบบคำพูด (สัปดาห์นี้/หน้า/ที่แล้ว, อีก N สัปดาห์, เดือนนี้/หน้า/ที่แล้ว, this/next/last week, this/next/last month) ให้ส่งวลีนั้นเป็น argument `week` ของ tool ตรงๆ — ระบบจะแปลงเป็น YW ให้เอง ห้ามเดา YW เอง
5. Earliest plannable week = YW {min_yw} (current +2 weeks) — exclude YW below this
6. Ava = available machines (Total − Used_N − Used_F). KG_Ava = available production capacity in kg (from KG_Ava_Display measure).
   - NEVER show raw column names or formulas to the user — e.g. "Used_N", "Used_F", "(Used_N + Used_F)", "Ava", "Total" must not appear in any response. Translate to plain words: Total → เครื่องทั้งหมด / total machines, Used_N → ใช้งานแผน Normal / in use (Normal), Used_F → ใช้งาน FQC / in use (FQC), Ava → เครื่องว่าง / available machines. Example: say "ใช้งานอยู่ **12 เครื่อง** (Normal 10 + FQC 2)" — never "(Used_N + Used_F) = 12". When Ava > 0, always include KG_Ava in the response (format as kg with comma separator, e.g. "**X,XXX kg**"). If KG_Ava is blank/empty for a row, omit it.
7. Be concise. Use bold (**text**) for key numbers. Use exactly ONE newline (\\n) between each bullet and between paragraphs. NEVER put multiple bullets on the same line. No blank lines (double newlines \\n\\n) anywhere in the response.
   - If tool result contains a line starting with [หมายเหตุ:...], you MUST include that warning in your response.
   - If tool result contains TOTAL_KP_WEIGHT=..., use that exact value for the total — never compute the sum yourself. Do NOT copy or show the [TOTAL_KP_WEIGHT=...] line in your response; it is for your internal use only.
8. When user greets (สวัสดี, hello, hi, etc.) — greeting คือคำทักทายล้วนๆ เท่านั้น:
   - Thai: "สวัสดีค่ะ น้อง I-SAVE Chatbot ค่ะพี่ๆ สามารถสอบถามข้อมูล หรือพิมพ์คำถามที่ต้องการได้เลยนะคะ น้องยินดีช่วยเหลือค่ะ"
   - English: "Hello! I'm I-SAVE Chatbot. Feel free to ask me anything about the I-SAVE system. I'm happy to help!"
   - คำถามเช่น "มีอะไรบ้าง", "ดูข้อมูล", "มีข้อมูลอะไร" ไม่ใช่การทักทาย → ให้ถือว่าถามภาพรวม I-SAVE แล้วตอบตามข้อ 12
   - Small-talk / courtesy (ขอบคุณ, ขอบใจ, thanks, thank you, โอเค, ok, เยี่ยม, ดีมาก, เก่งมาก, บาย, bye, ลาก่อน) — DO NOT call any tool and DO NOT use the rejection message in rule 9. Reply warmly and briefly, then invite the next question:
     - Thai (ขอบคุณ/ชม): "ยินดีค่ะ 😊 ถ้ามีอะไรให้ช่วยเพิ่มเติม สอบถามน้องได้เลยนะคะ"
     - Thai (ลา/บาย): "ขอบคุณค่ะ แล้วพบกันใหม่นะคะ 😊"
     - English (thanks/praise): "You're welcome! 😊 Feel free to ask if you need anything else."
     - English (goodbye): "Thank you! See you next time. 😊"
9. FIRST check: does the message relate to any of these I-SAVE topics?
   → items / item codes / item plan / แผน item / ข้อมูล item
   → machine groups / เครื่องจักร / เครื่องทอ / ข้อมูลเครื่อง
   → machine capacity / กำลังการผลิต / เครื่องว่าง / capacity
   → booking / การจอง / ข้อมูลการจอง
   → knit plan / แผนทอ / แผนการทอ / ข้อมูลแผนทอ / แผนการผลิต
   → KP Weight / week / YW / gauge
   → sales / ยอดของ sales / ยอดจอง / จัดอันดับ sales
   → export / ขอไฟล์ / ขอเป็น Excel / ดาวน์โหลดข้อมูล / กราฟ / ตาราง — คำขอ format/export ของข้อมูล I-SAVE ที่เพิ่งคุยกัน ถือเป็นหัวข้อ I-SAVE เสมอ (ดู rule 15/16) ห้าม reject
   If YES → ALWAYS call the appropriate tool. Even if no item/group/week is specified, still call the tool with no filter and give an overview per rule 12. NEVER reject an I-SAVE topic.
   If the message is clearly unrelated to I-SAVE (e.g. weather, cooking, today's date, general knowledge) → reply with this only (no extra text):
   - Thai: "ขออภัยค่ะ น้อง I-SAVE Chatbot ยังไม่สามารถตอบคำถามนี้ได้ในขณะนี้\nรบกวนพี่ๆ สอบถามเฉพาะข้อมูลที่เกี่ยวข้องกับระบบ I-SAVE หรือหัวข้อที่น้องรองรับนะคะ"
   - English: "Sorry, I'm unable to answer this question. Please ask only about I-SAVE system data or supported topics."
10. Week terminology: Thai responses → always use "สัปดาห์" (e.g. "สัปดาห์ที่ 22") — never "week 22" or "วีค 22". English responses → use "week" (e.g. "Week 22").
    Gauge terminology: always write "Gauge" or abbreviate as "G" (e.g. "24G", "28G") — never use "เกจ" in any language.
11. When multiple machine groups have available capacity, give a summary first:
    - Thai: "มีเครื่องว่างรวม **XX เครื่อง** ใน YY กลุ่ม (KG_Ava รวม **X,XXX kg**)" แล้วถามว่า "ต้องการดูรายละเอียดแต่ละกลุ่มเพิ่มเติมไหมคะ"
    - English: "Total **XX machines** available across YY groups (KG_Ava: **X,XXX kg**). Would you like details per group?"
    - Include KG_Ava only when KG_Ava column has values; omit the "(KG_Ava ...)" part if blank.
12. If no Item or group is specified, give an overview:
    - Thai: "พบ Item ในแผนทั้งหมด [จำนวน] รายการ อยู่ใน [จำนวน] กลุ่มเครื่อง เช่น [รายชื่อกลุ่มหลัก]\nKP_Weight รวม [ยอดรวม] kg\nต้องการดูรายละเอียดเพิ่มเติมไหมคะ เช่น Item, กลุ่มเครื่อง, Gauge หรือช่วงสัปดาห์"
    - English: "Found [count] items in the plan across [count] machine groups (e.g. [main groups]).\nTotal KP_Weight: [total] kg.\nWould you like more details by item, group, Gauge, or week range?"
13. Users often type only a partial item code — either a prefix (e.g. "F100114") or a substring from the middle/end of the code (e.g. "PI80" for "FD4BASPI80B0", "JZ63" for a code containing "JZ63"). Always pass whatever the user typed as item_code to get_item_plan; the tool handles prefix and substring matching automatically. If multiple items match, show a numbered list (1, 2, 3, 4) and ask which one the user wants before showing details.
14. When asked about a specific item, call get_item_plan first, then format the response using ACTUAL values from the tool result.
    CRITICAL: NEVER output literal bracket placeholders. Replace every placeholder with real data. If tool returns no data, say item was not found.
    Format template (replace ALL bracketed parts with real values from tool result):
    - Thai: "Item {{actual_item_code}} สามารถทอได้ที่กลุ่มเครื่อง {{actual_group_name}} โดยมีรายละเอียดแผนทอ ดังนี้\n1.สัปดาห์ที่ {{actual_YW}} จำนวน {{actual_KP_Weight}} kg\n...\nหากต้องการสอบถามเรื่องไหนเพิ่มเติม สามารถพิมพ์สอบถามได้เลยค่ะ"
    - English: "Item {{actual_item_code}} can be knitted at machine group {{actual_group_name}}. Knitting plan details:\n1. Week {{actual_YW}}: {{actual_KP_Weight}} kg\n...\nFeel free to ask if you need more information."
    - Display KP_Weight in kg with comma separator (e.g. 859,739 kg). Do NOT convert to tons.
    - If item spans multiple groups, list each group separately
15. Charts/graphs (กราฟ, แผนภูมิ, chart, plot, วาดกราฟ): ONLY when the user explicitly asks for one.
    - Step 1: call the data tool(s) first (get_item_plan ฯลฯ) to get REAL values. NEVER invent numbers.
    - Step 2: call render_chart with those real values. Use KP_Weight in kg directly (no conversion).
    - chart_type: 'line' = trend over weeks (แนวโน้มรายสัปดาห์), 'bar' = compare weeks/groups, 'pie'/'doughnut' = proportions/สัดส่วน. If user names a type, use it.
    - labels = x-axis (เช่น สัปดาห์ YW), each dataset.data must align 1:1 with labels.
    - After render_chart, write only a SHORT caption (1-2 lines). Do NOT re-list every number — the chart already shows them.
    - If there is no data to plot (tool returned not found), do NOT call render_chart; reply that there is no data to chart.
16. Excel export (export, ขอเป็นไฟล์, ดาวน์โหลดเป็น Excel/xlsx): ONLY when the user explicitly asks for a file/export.
    - SCOPE (สำคัญที่สุด): export ข้อมูลชุดเดียวกับที่เพิ่งคุย/แสดงในรอบก่อนเป๊ะๆ — ใช้ filter เดิมทั้งหมด (group/week/item/gauge). เช่น เพิ่งคุยเรื่อง capacity สัปดาห์ที่ 26 → ต้องส่ง week นั้นให้ data tool ด้วย ห้ามดึงทุกสัปดาห์
    - ถ้าบริบทก่อนหน้าไม่มี filter เลย และการ export ทั้งหมดจะใหญ่มาก (ทุกสัปดาห์ × ทุกกลุ่ม) → ถาม user ก่อนว่าต้องการสัปดาห์ไหน/กลุ่มไหน ห้าม export ทั้ง dataset เด็ดขาด
    - HOW: call export_excel ONCE with source_tool (get_item_plan/get_machine_capacity/get_booking) + those filters — the system fetches and fills ALL rows server-side.
    - NEVER pass rows yourself for tool data (arguments WILL be truncated when data is large and the export will fail). columns+rows mode is ONLY for small data that did not come from a data tool.
    - ตั้ง filename ให้สื่อ scope เช่น capacity_skp_202626 (ห้ามใช้คำว่า all_weeks ถ้าไม่ได้ export ทุกสัปดาห์จริง)
    - The tool result contains a markdown download link — you MUST include that link line in your reply EXACTLY as given (do not modify the URL).
17. When your tool results contain 3 or more data rows to present, follow this exact 3-step order:
    Step 1 — write a summary block BEFORE calling render_table:
      📌 สรุป[กลุ่ม/ตัวกรองที่ใช้] [สัปดาห์ถ้ามี]
      KP Weight รวม: [ค่าจาก TOTAL_KP_WEIGHT] kg
      จำนวน Item: [จำนวน] รายการ
      Item ที่มีน้ำหนักมากที่สุด: (Top 3 เรียงมาก→น้อย, format: "ItemCode — X,XXX.XX kg" แต่ละรายการขึ้นบรรทัดใหม่)
      (ถ้าไม่มี group/week filter ให้ปรับหัวข้อตามข้อมูลจริง เช่น "📌 สรุป Item ทั้งหมดสัปดาห์ที่ XX")
    Step 2 — call render_table with ALL rows (ตารางเต็ม)
    Step 3 — write 1–2 short follow-up suggestions (เช่น ดูกรองเพิ่ม, ดูสัปดาห์อื่น)
    - MANDATORY for EVERY response that has 3+ rows — even if a previous turn already showed a table for the same data source. NEVER substitute a numbered list, bullet list, or inline text for render_table.
    - columns: array of header label strings, e.g. ["Item", "กลุ่ม", "KP Weight (kg)", "สัปดาห์"]
    - rows: array of arrays — one sub-array per data row, values in the same order as columns
    - LIMIT: maximum 50 rows per table call. If data has more than 50 rows, show only the first 50 and mention the total count in your caption.
    - Number columns (KP Weight, Total, Used, Ava, etc.): use raw numeric values (int/float) — the table auto-formats with commas
    - Code/ID columns (Item codes, YW week codes like 202626, group names): MUST be passed as strings (e.g. "202626" not 202626) — never pass week codes as numbers
    - Call the relevant data tool(s) first to get REAL values, then call render_table with those values
    - Only use render_chart (not render_table) when the user explicitly asks for a graph/chart
18. I-SAVE system how-to questions — answer from the knowledge below, do NOT call any tool:
    Q: เปลี่ยนรหัสผ่านยังไง / วิธีเปลี่ยนรหัสผ่าน / change password
    A (Thai): กดที่ชื่อผู้ใช้มุมบนขวา → เลือก "เปลี่ยนรหัสผ่าน" → กรอกรหัสผ่านเดิม และรหัสผ่านใหม่ (อย่างน้อย 6 ตัวอักษร) → กด "ยืนยัน"
    A (English): Click your username at the top-right → select "เปลี่ยนรหัสผ่าน" (Change Password) → enter your current password and new password (min 6 characters) → click "ยืนยัน" (Confirm).

    Q: ลืมรหัสผ่าน / เข้าระบบไม่ได้ / forgot password / reset password
    A (Thai): ที่หน้า Login กดลิงก์ "ลืมรหัสผ่าน" → กรอกชื่อผู้ใช้และอีเมลที่ลงทะเบียนไว้ให้ตรงกัน → ตั้งรหัสผ่านใหม่ได้เลย
    A (English): On the Login page, click "ลืมรหัสผ่าน" (Forgot Password) → enter your username and registered email → set a new password.

    Q: ติดต่อแอดมินยังไง / แจ้งปัญหา / contact admin / report issue
    A (Thai): กดปุ่ม "ติดต่อ Admin" ที่แถบซ้าย → พิมพ์อธิบายปัญหา → กด "ส่งข้อความ" แอดมินจะตอบกลับในระบบ ดูประวัติได้ที่แท็บ "ประวัติ"
    A (English): Click "ติดต่อ Admin" in the left sidebar → describe your issue → click "ส่งข้อความ" (Send). Admin will reply in the system; view history under the "ประวัติ" tab.

    Q: ออกจากระบบยังไง / logout / sign out
    A (Thai): กดปุ่ม "ออกจากระบบ" สีแดงมุมบนขวา
    A (English): Click the red "ออกจากระบบ" (Logout) button at the top-right."""

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "get_item_plan",
            "description": (
                "ดึงข้อมูล Item Plan จาก BookingMaster: "
                "Item=รหัสสินค้า, Group=กลุ่มเครื่อง, KP_Weight=น้ำหนักที่วางแผน (kg), YW=สัปดาห์. "
                "ใช้ตอบคำถามเกี่ยวกับ item plan, กลุ่มเครื่องของ item, KP Weight, สัปดาห์ที่จะทอ"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "item_code": {
                        "type": "string",
                        "description": "รหัส item เช่น F100114/10A0 ถ้าไม่ระบุจะคืนทุก item"
                    },
                    "group": {
                        "type": "string",
                        "description": "กรองเฉพาะกลุ่มเครื่อง เช่น SKP, SKPLE, SKPTA"
                    },
                    "week": {
                        "type": "string",
                        "description": "สัปดาห์ — ส่งได้ทั้งเลข (week22, wk22, 202622) และคำพูดสัมพัทธ์ตรงๆ เช่น 'สัปดาห์นี้/หน้า/ที่แล้ว', 'อีก 3 สัปดาห์', 'เดือนนี้/หน้า/ที่แล้ว' (ระบบคำนวณ YW ให้เอง). ไม่ระบุ = ทุกสัปดาห์"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_machine_capacity",
            "description": (
                "ดึงข้อมูลกำลังการผลิต Table_MC: "
                "YW=สัปดาห์, Group=กลุ่มเครื่อง, Guage=gauge (column name in DB is misspelled), "
                "Total=เครื่องทั้งหมด, Used_N=ใช้ Normal, Used_F=ใช้ FQC, Ava=เครื่องว่าง, "
                "KG_Ava=กำลังการผลิตที่ว่างในหน่วย kg (มีค่าเฉพาะแถวที่ Ava > 0). "
                "ใช้ตอบคำถามเรื่องเครื่องว่าง, กำลังการผลิต, เครื่องทั้งหมดในกลุ่ม"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "group": {
                        "type": "string",
                        "description": "กลุ่มเครื่อง เช่น SKP, SKPLE, SKPTA ถ้าไม่ระบุจะคืนทุกกลุ่ม"
                    },
                    "week": {
                        "type": "string",
                        "description": "สัปดาห์ — ส่งได้ทั้งเลข (week22, wk22, 202622) และคำพูดสัมพัทธ์ตรงๆ เช่น 'สัปดาห์นี้/หน้า/ที่แล้ว', 'อีก 3 สัปดาห์', 'เดือนนี้/หน้า/ที่แล้ว' (ระบบคำนวณ YW ให้เอง). ไม่ระบุ = ทุกสัปดาห์"
                    },
                    "gauge": {
                        "type": "string",
                        "description": "Gauge ของเครื่อง เช่น 20, 24, 28 ถ้าไม่ระบุจะคืนทุก gauge"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_booking",
            "description": (
                "ดึงข้อมูล Booking Master: "
                "YW=สัปดาห์, MC_GROUP=กลุ่มเครื่อง, Used=จำนวนเครื่องที่จองแล้ว. "
                "ใช้ตอบคำถามเรื่องการจองเครื่อง"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "group": {
                        "type": "string",
                        "description": "กลุ่มเครื่อง ถ้าไม่ระบุจะคืนทุกกลุ่ม"
                    },
                    "week": {
                        "type": "string",
                        "description": "สัปดาห์ — ส่งได้ทั้งเลข (week22, 202622) และคำพูดสัมพัทธ์ตรงๆ เช่น 'สัปดาห์นี้/หน้า/ที่แล้ว', 'อีก 3 สัปดาห์', 'เดือนนี้/หน้า/ที่แล้ว'. ไม่ระบุ = ทุกสัปดาห์"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_knit_plan",
            "description": (
                "ดึงแผนการทอ (Knit Plan): "
                "Item=รหัสสินค้า, Group=กลุ่มเครื่อง, KP_Weight=น้ำหนัก (kg), YW=สัปดาห์. "
                "ใช้ตอบคำถามเรื่องแผนการทอ รายการสินค้าในแต่ละสัปดาห์"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "week": {
                        "type": "string",
                        "description": "สัปดาห์ — ส่งได้ทั้งเลข (week22, 202622) และคำพูดสัมพัทธ์ตรงๆ เช่น 'สัปดาห์นี้/หน้า/ที่แล้ว', 'อีก 3 สัปดาห์', 'เดือนนี้/หน้า/ที่แล้ว'. ไม่ระบุ = ทุกสัปดาห์"
                    },
                    "group": {
                        "type": "string",
                        "description": "กลุ่มเครื่อง ถ้าไม่ระบุจะคืนทุกกลุ่ม"
                    },
                    "item_code": {
                        "type": "string",
                        "description": "รหัส item ถ้าไม่ระบุจะคืนทุก item"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "render_table",
            "description": (
                "แสดงข้อมูลหลาย row เป็นตาราง interactive ที่ sort/filter/export CSV ได้ "
                "เรียกใช้เมื่อมีข้อมูล 3+ แถว (เช่น item หลายรายการ, หลายกลุ่มเครื่อง, หลายสัปดาห์). "
                "ต้องเรียก tool ดึงข้อมูล (get_item_plan ฯลฯ) ก่อน แล้วนำค่าจริงมาใส่. "
                "ห้ามเรียกเมื่อ user ขอกราฟ — ให้ใช้ render_chart แทน"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "title": {
                        "type": "string",
                        "description": "หัวข้อตาราง เช่น 'Item Plan กลุ่ม SKP สัปดาห์ที่ 25'"
                    },
                    "columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "ชื่อ header แต่ละคอลัมน์ เช่น ['Item', 'กลุ่ม', 'KP Weight (kg)', 'สัปดาห์']"
                    },
                    "rows": {
                        "type": "array",
                        "description": "ข้อมูลแต่ละแถว เรียงตรงกับ columns",
                        "items": {
                            "type": "array",
                            "items": {"type": ["string", "number", "null"]}
                        }
                    }
                },
                "required": ["columns", "rows"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "render_chart",
            "description": (
                "วาดกราฟ/แผนภูมิให้ผู้ใช้ เรียกใช้เมื่อผู้ใช้ขอกราฟ/แผนภูมิ/chart/plot/วาดกราฟ เท่านั้น. "
                "ต้องเรียก tool ดึงข้อมูล (get_item_plan ฯลฯ) ให้ได้ค่าจริงก่อน แล้วนำตัวเลขจริงมาใส่. "
                "เลือก chart_type: 'line'=แนวโน้มตามสัปดาห์, 'bar'=เปรียบเทียบ (สัปดาห์/กลุ่ม), "
                "'pie'/'doughnut'=สัดส่วน. labels และ data ต้องยาวเท่ากัน. ใช้ KP_Weight ในหน่วย kg ตรงๆ ไม่ต้องแปลง"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "chart_type": {
                        "type": "string",
                        "enum": ["bar", "line", "pie", "doughnut"],
                        "description": "ชนิดกราฟ"
                    },
                    "title": {
                        "type": "string",
                        "description": "หัวข้อกราฟ เช่น 'แผนทอ F100114/10A0 รายสัปดาห์ (kg)'"
                    },
                    "labels": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "ป้ายแกน X เช่น ['202615','202617','202618']"
                    },
                    "datasets": {
                        "type": "array",
                        "description": "ชุดข้อมูล แต่ละชุด = {label, data}. ปกติมีชุดเดียว",
                        "items": {
                            "type": "object",
                            "properties": {
                                "label": {"type": "string", "description": "ชื่อชุดข้อมูล เช่น 'KP Weight (kg)'"},
                                "data": {
                                    "type": "array",
                                    "items": {"type": "number"},
                                    "description": "ค่าตัวเลข เรียงตรงกับ labels"
                                }
                            },
                            "required": ["label", "data"]
                        }
                    },
                    "y_label": {
                        "type": "string",
                        "description": "ชื่อแกน Y เช่น 'kg' (ไม่บังคับ)"
                    }
                },
                "required": ["chart_type", "labels", "datasets"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "suggest_week",
            "description": (
                "หาสัปดาห์ (YW) เร็วที่สุดที่มีเครื่องว่างให้วางแผน โดยเริ่มจาก min plannable week "
                "(current week + 2) ให้อัตโนมัติ — ห้ามเรียก get_machine_capacity มาไล่หาเอง. "
                "ใช้เมื่อ user ถาม 'ลงเครื่องได้เร็วสุด week ไหน', 'ควรวางแผน week ไหน', 'มีที่ว่างเมื่อไหร่'. "
                "คืนรายการสัปดาห์ที่ Ava >= machines_needed เรียงจากเร็วไปช้า"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "group": {
                        "type": "string",
                        "description": "กลุ่มเครื่อง เช่น SKP, SKPLE, SKPTA — ควรระบุถ้า user บอก"
                    },
                    "gauge": {
                        "type": "string",
                        "description": "Gauge ของเครื่อง เช่น 20, 24, 28 ถ้าไม่ระบุจะรวมทุก gauge"
                    },
                    "machines_needed": {
                        "type": "number",
                        "description": "จำนวนเครื่องว่างขั้นต่ำที่ต้องการ (default 1)"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "compare_weeks",
            "description": (
                "เปรียบเทียบกำลังการผลิตระหว่าง 2 สัปดาห์ คืนตารางต่อกลุ่มเครื่อง: "
                "Total และ Ava ของแต่ละสัปดาห์ พร้อม Ava_Delta (week_b − week_a) และ Ava_Pct "
                "คำนวณให้แล้ว — ใช้ค่าจากผลลัพธ์ตรงๆ ห้ามคำนวณเอง. "
                "ใช้เมื่อ user ถามเทียบ capacity ระหว่างสัปดาห์ เช่น 'week นี้เทียบกับ week ที่แล้ว'"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "week_a": {
                        "type": "string",
                        "description": "สัปดาห์ฐาน (base) — เลข (202622) หรือคำพูดสัมพัทธ์ ('สัปดาห์นี้', 'สัปดาห์ที่แล้ว')"
                    },
                    "week_b": {
                        "type": "string",
                        "description": "สัปดาห์ที่นำมาเทียบ — เลข (202624) หรือคำพูดสัมพัทธ์ ('สัปดาห์หน้า', 'อีก 2 สัปดาห์')"
                    },
                    "group": {
                        "type": "string",
                        "description": "กรองเฉพาะกลุ่มเครื่อง ถ้าไม่ระบุจะเทียบทุกกลุ่ม"
                    }
                },
                "required": ["week_a", "week_b"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "item_capability",
            "description": (
                "บอกว่า item ทอได้ที่กลุ่มเครื่อง/SubGroup ไหนบ้าง (จาก Table_Item master) "
                "พร้อมบอกว่ากลุ่มไหนมีแผนปัจจุบันอยู่แล้ว และแต่ละกลุ่มมีเครื่องว่างเร็วสุดสัปดาห์ไหน. "
                "ใช้เมื่อ user ถาม 'item นี้ทอได้ที่กลุ่มไหนบ้าง', 'ย้ายไปกลุ่มอื่นได้ไหม', 'มีกลุ่มสำรองไหม'. "
                "ต่างจาก get_item_plan ที่บอกแค่กลุ่มที่มีแผนอยู่แล้ว"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "item_code": {
                        "type": "string",
                        "description": "รหัส item (รองรับ prefix/substring)"
                    },
                    "week": {
                        "type": "string",
                        "description": "ถ้าระบุ จะแสดงเครื่องว่างของแต่ละกลุ่ม ณ สัปดาห์นั้นแทน next free week"
                    }
                },
                "required": ["item_code"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_sales_summary",
            "description": (
                "ยอดจอง (kg) และจำนวน item ต่อ sales จาก BookingMaster. "
                "ใช้เมื่อ user ถามยอดของ sales คนใดคนหนึ่ง หรือจัดอันดับ sales. "
                "รองรับกรองช่วงเวลา (สัปดาห์/เดือน) ผ่าน parameter week"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "sales_name": {
                        "type": "string",
                        "description": "ชื่อ sales (รองรับพิมพ์บางส่วน) ไม่ระบุ = จัดอันดับทุกคน"
                    },
                    "week": {
                        "type": "string",
                        "description": "ช่วงเวลา — เลข (202626) หรือคำพูดสัมพัทธ์ ('สัปดาห์นี้', 'เดือนนี้', 'เดือนหน้า') ไม่ระบุ = ยอดรวมทั้งหมด"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "group_utilization",
            "description": (
                "จัดอันดับการใช้งานกลุ่มเครื่อง: Total, Used, Ava, Used% (คำนวณให้แล้ว — ห้ามคำนวณเอง), "
                "KG_Ava และสัปดาห์ที่วิกฤตที่สุดของแต่ละกลุ่ม. "
                "ใช้เมื่อ user ถาม 'กลุ่มไหนแน่น/ตึงสุด', 'กลุ่มไหนว่างสุด', 'ภาพรวมการใช้เครื่อง'"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "week": {
                        "type": "string",
                        "description": "ช่วงเวลา — เลข (202626) หรือคำพูด ('สัปดาห์นี้', 'เดือนหน้า') ไม่ระบุ = 4 สัปดาห์ข้างหน้า"
                    },
                    "weeks_ahead": {
                        "type": "number",
                        "description": "จำนวนสัปดาห์ข้างหน้าที่จะวิเคราะห์ (default 4, สูงสุด 12) — ใช้เมื่อไม่ระบุ week"
                    },
                    "sort_by": {
                        "type": "string",
                        "enum": ["used_pct", "available"],
                        "description": "used_pct = กลุ่มตึงสุดก่อน (default), available = กลุ่มว่างมากสุดก่อน"
                    }
                },
                "required": []
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "analyze_plan_impact",
            "description": (
                "วิเคราะห์ผลกระทบถ้าถอด/เลื่อน/ย้ายแผนของ item: บอกว่ากลุ่มเครื่องที่ item ใช้แน่นแค่ไหน "
                "ถ้าถอดออก Ava จะเพิ่มประมาณเท่าไร และมี item ใดใช้กลุ่ม+สัปดาห์เดียวกันบ้าง (ผู้ที่ได้/เสียประโยชน์). "
                "ใช้เมื่อ user ถาม 'ถ้าถอดแผนนี้กระทบ item อื่นไหม', 'เลื่อนแผนได้ไหม', 'item นี้ใช้เครื่องร่วมกับใคร'"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "item_code": {
                        "type": "string",
                        "description": "รหัส item ที่จะวิเคราะห์ เช่น F100114/10A0 (รองรับ prefix/substring)"
                    },
                    "week": {
                        "type": "string",
                        "description": "เจาะจงสัปดาห์ — เลข (202626) หรือคำพูดสัมพัทธ์ ไม่ระบุ = วิเคราะห์ทุกสัปดาห์ที่มีแผน"
                    },
                    "group": {
                        "type": "string",
                        "description": "เจาะจงกลุ่มเครื่อง ถ้า item อยู่หลายกลุ่ม"
                    }
                },
                "required": ["item_code"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "export_excel",
            "description": (
                "สร้างไฟล์ Excel (.xlsx) ให้ user ดาวน์โหลด เรียกเฉพาะเมื่อ user ขอ export/ไฟล์/Excel/ดาวน์โหลดข้อมูล. "
                "วิธีที่ถูกต้อง: ส่ง source_tool + filter (item_code/group/week/gauge) ตาม scope ที่คุยกัน — "
                "ระบบจะดึงข้อมูลและใส่ทุกแถวให้เอง ห้ามส่ง rows เอง (arguments จะ truncate เมื่อข้อมูลเยอะ). "
                "ผลลัพธ์มีลิงก์ดาวน์โหลด — ต้องใส่ลิงก์นั้นในคำตอบเสมอ"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "source_tool": {
                        "type": "string",
                        "enum": ["get_item_plan", "get_knit_plan", "get_machine_capacity", "get_booking"],
                        "description": "data tool ที่จะดึงข้อมูลมา export โดยตรง (แนะนำเสมอ) — ใช้คู่กับ filter ด้านล่าง ไม่ต้องส่ง columns/rows"
                    },
                    "item_code": {
                        "type": "string",
                        "description": "filter: รหัส item (สำหรับ get_item_plan/get_knit_plan)"
                    },
                    "group": {
                        "type": "string",
                        "description": "filter: กลุ่มเครื่อง เช่น SKP"
                    },
                    "week": {
                        "type": "string",
                        "description": "filter: สัปดาห์ — เลข (202626) หรือคำพูดสัมพัทธ์ ('สัปดาห์นี้') ตาม scope ที่คุยกัน"
                    },
                    "gauge": {
                        "type": "string",
                        "description": "filter: gauge (สำหรับ get_machine_capacity)"
                    },
                    "title": {
                        "type": "string",
                        "description": "หัวข้อรายงานในไฟล์ เช่น 'แผนทอ SKP สัปดาห์ 202626'"
                    },
                    "filename": {
                        "type": "string",
                        "description": "ชื่อไฟล์สื่อ scope (a-z, 0-9, -, _ ไม่ต้องใส่ .xlsx) เช่น item_plan_202626"
                    },
                    "columns": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "โหมด manual เท่านั้น (ข้อมูลไม่ได้มาจาก data tool): หัวตาราง"
                    },
                    "rows": {
                        "type": "array",
                        "description": "โหมด manual เท่านั้น: ข้อมูลแต่ละแถว — ห้ามใช้ถ้ามี source_tool",
                        "items": {"type": "array"}
                    }
                },
                "required": []
            }
        }
    }
]


class ResponseProcessor:
    def __init__(self):
        self._openai = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        self._openai_async = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
        self._model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

    def _build_history_messages(self, conversation_history: list, limit: int = 12) -> list:
        # conversation_history มาจาก get_conversation_messages ที่ดึง "ก่อน" บันทึก message ปัจจุบัน
        # (chatbot_app.send_message) → element สุดท้ายคือคำตอบ assistant ของรอบก่อน ต้องเก็บไว้
        # ห้าม slice [:-1] เพราะจะตัดคำตอบรอบก่อนทิ้ง ทำให้โมเดลคิดว่าคำถามเก่ายังไม่ถูกตอบแล้วตอบซ้ำ
        messages = []
        recent = conversation_history[-limit:]
        for msg in recent:
            role = msg.get('sender') or msg.get('role') or msg.get('message_type', '')
            content = msg.get('content') or msg.get('message', '')
            if not content:
                continue
            if role == 'user':
                messages.append({"role": "user", "content": content})
            elif role in ('assistant', 'ai'):
                messages.append({"role": "assistant", "content": content})
        return messages

    # ---- Tool implementations ----

    def _tool_get_item_plan(self, item_code: str = None, group: str = None, week: str = None) -> str:
        cached, ready = get_data_cache().get('query_item')
        if not ready or not cached:
            return "ข้อมูล Item Plan ยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        data = cached.get('data', '')
        if not data:
            return "ไม่มีข้อมูล Item Plan"

        lines = data.splitlines()
        header = lines[0]
        # columns: Item,Group,KP_Weight,YW

        yw_filter = None
        if week:
            yws = _week_keyword_to_yw(week)
            if not yws and re.match(r'^\d{6}$', week.strip()):
                yws = [week.strip()]
            yw_filter = set(yws) if yws else None

        result = []
        exact_match = []
        prefix_match = []
        substr_match = []
        ic = item_code.strip().upper() if item_code else None
        for line in lines[1:]:
            cols = line.split(',')
            if not cols:
                continue
            item_col = cols[0].strip().upper() if len(cols) > 0 else ''
            group_col = cols[1].strip().lower() if len(cols) > 1 else ''
            yw_col = cols[3].strip() if len(cols) > 3 else ''

            if group and not _group_matches(group_col, [group.lower()]):
                continue
            if yw_filter and yw_col not in yw_filter:
                continue
            if ic:
                if item_col == ic:
                    exact_match.append(line)
                elif item_col.startswith(ic):
                    prefix_match.append(line)
                elif ic in item_col:
                    substr_match.append(line)
                continue
            result.append(line)

        if ic:
            result = exact_match or prefix_match or substr_match

        if result:
            min_yw = _min_plannable_yw()
            if yw_filter and max(yw_filter) < min_yw:
                yw_show = ', '.join(sorted(yw_filter))
                note = f"[หมายเหตุ: YW {yw_show} ผ่านมาแล้ว สัปดาห์เร็วที่สุดที่วางแผนได้ = YW {min_yw}]\n"
            else:
                past = [l for l in result if len(l.split(',')) > 3 and l.split(',')[3].strip() < min_yw]
                note = f"[หมายเหตุ: มี {len(past)} รายการที่อยู่ใน YW ที่ผ่านมาแล้ว (ก่อน YW {min_yw})]\n" if past else ""
            total_kg = 0.0
            for l in result:
                cols = l.split(',')
                if len(cols) > 2:
                    try:
                        total_kg += float(cols[2].strip())
                    except ValueError:
                        pass
            footer = f"[TOTAL_KP_WEIGHT={total_kg:.2f} kg — ใช้ค่านี้เท่านั้น ห้ามคำนวณเอง]\n"
            return header + '\n' + '\n'.join(result) + '\n' + footer + note
        if item_code:
            return f"ไม่พบ item {item_code} ใน Item Plan"
        return "ไม่พบข้อมูลที่ตรงกับเงื่อนไข"

    def _tool_get_machine_capacity(self, group: str = None, week: str = None, gauge: str = None) -> str:
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return "ข้อมูลกำลังการผลิตยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        data = cached.get('data', {})
        csv_text = data.get('mc', '') if isinstance(data, dict) else ''
        if not csv_text:
            return "ไม่มีข้อมูล Machine Capacity"

        # Build KG_Ava lookup: (yw, group) -> kg_ava
        kg_ava_lookup: dict = {}
        kg_ava_csv = data.get('kg_ava', '') if isinstance(data, dict) else ''
        if kg_ava_csv:
            for line in kg_ava_csv.splitlines()[1:]:
                parts = line.split(',')
                if len(parts) >= 3:
                    try:
                        kg_ava_lookup[(parts[0].strip(), parts[1].strip().lower())] = float(parts[2].strip())
                    except ValueError:
                        pass

        lines = csv_text.splitlines()
        # columns: YW,Group,Guage,Total,Used_N,Used_F,Ava

        yw_filter = None
        if week:
            yws = _week_keyword_to_yw(week)
            if not yws and re.match(r'^\d{6}$', week.strip()):
                yws = [week.strip()]
            yw_filter = set(yws) if yws else None

        result = []
        for line in lines[1:]:
            cols = line.split(',')
            if not cols:
                continue
            yw_col = cols[0].strip() if len(cols) > 0 else ''
            group_col = cols[1].strip().lower() if len(cols) > 1 else ''
            gauge_col = cols[2].strip() if len(cols) > 2 else ''
            ava_col = cols[6].strip() if len(cols) > 6 else ''

            if yw_filter and yw_col not in yw_filter:
                continue
            if group and not _group_matches(group_col, [group.lower()]):
                continue
            if gauge and gauge_col != gauge:
                continue

            # KG_Ava เป็นค่ารวมระดับกลุ่ม ไม่ break down ตาม gauge
            # ถ้า filter ด้วย gauge ให้ข้าม KG_Ava เพื่อไม่ให้เข้าใจผิด
            if gauge:
                result.append(line)
            else:
                try:
                    ava_val = float(ava_col) if ava_col else 0.0
                except ValueError:
                    ava_val = 0.0
                kg_ava_val = kg_ava_lookup.get((yw_col, group_col))
                if kg_ava_val is not None and ava_val > 0:
                    result.append(f"{line},{kg_ava_val:.2f}")
                else:
                    result.append(f"{line},")

        if result:
            min_yw = _min_plannable_yw()
            note = f"[หมายเหตุ: week เร็วที่สุดที่วางแผนได้ = YW {min_yw}]\n"
            header = "YW,Group,Guage,Total,Used_N,Used_F,Ava" if gauge else "YW,Group,Guage,Total,Used_N,Used_F,Ava,KG_Ava"
            return header + '\n' + '\n'.join(result) + '\n' + note
        return f"ไม่พบข้อมูล Machine Capacity (group={group}, week={week}, gauge={gauge})"

    def _tool_get_booking(self, group: str = None, week: str = None) -> str:
        cached, ready = get_data_cache().get('query_booking')
        if not ready or not cached:
            return "ข้อมูล Booking ยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        data = cached.get('data', '')
        if not data:
            return "ไม่มีข้อมูล Booking"

        lines = data.splitlines()
        header = lines[0]
        # columns: YW,MC_GROUP,Used

        yw_filter = None
        if week:
            yws = _week_keyword_to_yw(week)
            if not yws and re.match(r'^\d{6}$', week.strip()):
                yws = [week.strip()]
            yw_filter = set(yws) if yws else None

        result = []
        for line in lines[1:]:
            cols = line.split(',')
            if not cols:
                continue
            yw_col = cols[0].strip() if len(cols) > 0 else ''
            group_col = cols[1].strip().lower() if len(cols) > 1 else ''

            if yw_filter and yw_col not in yw_filter:
                continue
            if group and not _group_matches(group_col, [group.lower()]):
                continue
            result.append(line)

        if result:
            return header + '\n' + '\n'.join(result)
        return f"ไม่พบข้อมูล Booking (group={group}, week={week})"

    def _tool_get_knit_plan(self, week: str = None, group: str = None, item_code: str = None) -> str:
        return self._tool_get_item_plan(item_code=item_code, group=group, week=week)

    @staticmethod
    def _parse_week_arg(week: str):
        """แปลง argument week (เลข 6 หลัก หรือคำพูดสัมพัทธ์) เป็น YW เดียว — None ถ้าแปลงไม่ได้"""
        if not week:
            return None
        yws = _week_keyword_to_yw(week)
        if not yws and re.match(r'^\d{6}$', week.strip()):
            yws = [week.strip()]
        return yws[0] if yws else None

    @staticmethod
    def _kg_ava_lookup(data: dict) -> dict:
        """สร้าง lookup (yw, group_lower) -> kg_ava จาก cache payload"""
        lookup: dict = {}
        kg_csv = data.get('kg_ava', '') if isinstance(data, dict) else ''
        for line in kg_csv.splitlines()[1:] if kg_csv else []:
            parts = line.split(',')
            if len(parts) >= 3:
                try:
                    lookup[(parts[0].strip(), parts[1].strip().lower())] = float(parts[2].strip())
                except ValueError:
                    pass
        return lookup

    def _tool_suggest_week(self, group: str = None, gauge: str = None, machines_needed=None, limit: int = 5) -> str:
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return "ข้อมูลกำลังการผลิตยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        data = cached.get('data', {})
        csv_text = data.get('mc', '') if isinstance(data, dict) else ''
        if not csv_text:
            return "ไม่มีข้อมูล Machine Capacity"

        try:
            need = float(machines_needed) if machines_needed is not None else 1.0
        except (TypeError, ValueError):
            need = 1.0
        if need < 1:
            need = 1.0

        kg_lookup = self._kg_ava_lookup(data)
        min_yw = _min_plannable_yw()
        gauge_s = str(gauge).strip() if gauge else None

        # รวม Ava ต่อ YW (เฉพาะ YW >= min plannable) + จดกลุ่มที่มีเครื่องว่าง
        per_yw: dict = {}
        for line in csv_text.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            yw = cols[0].strip()
            if yw < min_yw:
                continue
            group_col = cols[1].strip()
            if group and not _group_matches(group_col.lower(), [group.lower()]):
                continue
            if gauge_s and cols[2].strip() != gauge_s:
                continue
            try:
                ava = float(cols[6].strip() or 0)
            except ValueError:
                ava = 0.0
            rec = per_yw.setdefault(yw, {'ava': 0.0, 'groups': set()})
            rec['ava'] += ava
            if ava > 0:
                rec['groups'].add(group_col)

        candidates = sorted(yw for yw, rec in per_yw.items() if rec['ava'] >= need)
        if not candidates:
            return (
                f"ไม่พบสัปดาห์ที่มีเครื่องว่าง >= {need:g} เครื่อง "
                f"(group={group or 'ทุกกลุ่ม'}, gauge={gauge_s or 'ทุก gauge'}) "
                f"ตั้งแต่ YW {min_yw} เป็นต้นไปในช่วงข้อมูลที่มี"
            )

        # KG_Ava เป็นค่าระดับกลุ่ม — รวมได้เฉพาะตอนไม่ได้ filter gauge (convention เดียวกับ get_machine_capacity)
        lines_out = ["YW,Ava,KG_Ava,Groups"]
        for yw in candidates[:limit]:
            rec = per_yw[yw]
            kg = ''
            if not gauge_s:
                kg_sum = sum(
                    kg_lookup.get((yw, g.lower()), 0.0) for g in rec['groups']
                )
                if kg_sum > 0:
                    kg = f"{kg_sum:.2f}"
            groups_str = '; '.join(sorted(rec['groups']))
            lines_out.append(f"{yw},{rec['ava']:g},{kg},{groups_str}")

        note = (
            f"[คำแนะนำ: สัปดาห์เร็วที่สุดที่ลงได้คือ YW {candidates[0]} "
            f"(ต้องการ {need:g} เครื่อง). กฎ min plannable week = current+2 → YW {min_yw} ระบบกรองให้แล้ว]"
        )
        return '\n'.join(lines_out) + '\n' + note

    def _tool_compare_weeks(self, week_a: str = None, week_b: str = None, group: str = None) -> str:
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return "ข้อมูลกำลังการผลิตยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        data = cached.get('data', {})
        csv_text = data.get('mc', '') if isinstance(data, dict) else ''
        if not csv_text:
            return "ไม่มีข้อมูล Machine Capacity"

        yw_a = self._parse_week_arg(week_a)
        yw_b = self._parse_week_arg(week_b)
        if not yw_a or not yw_b:
            return f"ไม่เข้าใจสัปดาห์ที่ระบุ (week_a={week_a}, week_b={week_b}) — ส่งเป็นเลข YYYYWW เช่น 202624 หรือคำพูดเช่น 'สัปดาห์นี้'"
        if yw_a == yw_b:
            return f"week_a และ week_b เป็นสัปดาห์เดียวกัน (YW {yw_a}) — ไม่มีอะไรให้เปรียบเทียบ"

        # รวม Total/Ava ต่อ (group) ของแต่ละ week
        per_group: dict = {}
        for line in csv_text.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            yw = cols[0].strip()
            if yw not in (yw_a, yw_b):
                continue
            group_col = cols[1].strip()
            if group and not _group_matches(group_col.lower(), [group.lower()]):
                continue
            try:
                total = float(cols[3].strip() or 0)
                ava = float(cols[6].strip() or 0)
            except ValueError:
                continue
            slot = 'a' if yw == yw_a else 'b'
            rec = per_group.setdefault(group_col, {'a': [0.0, 0.0], 'b': [0.0, 0.0]})
            rec[slot][0] += total
            rec[slot][1] += ava

        if not per_group:
            return f"ไม่พบข้อมูลของ YW {yw_a} หรือ YW {yw_b} (group={group or 'ทุกกลุ่ม'})"

        header = f"Group,Total_{yw_a},Ava_{yw_a},Total_{yw_b},Ava_{yw_b},Ava_Delta,Ava_Pct"
        lines_out = [header]
        sum_a = [0.0, 0.0]
        sum_b = [0.0, 0.0]

        def _fmt_row(name, a, b):
            delta = b[1] - a[1]
            pct = f"{(delta / a[1] * 100):.1f}%" if a[1] else ''
            return f"{name},{a[0]:g},{a[1]:g},{b[0]:g},{b[1]:g},{delta:+g},{pct}"

        for gname in sorted(per_group):
            rec = per_group[gname]
            lines_out.append(_fmt_row(gname, rec['a'], rec['b']))
            sum_a[0] += rec['a'][0]
            sum_a[1] += rec['a'][1]
            sum_b[0] += rec['b'][0]
            sum_b[1] += rec['b'][1]
        if len(per_group) > 1:
            lines_out.append(_fmt_row('(รวม)', sum_a, sum_b))

        note = f"[Ava_Delta = YW {yw_b} − YW {yw_a}; ใช้ค่า Delta/Pct จากตารางนี้เท่านั้น ห้ามคำนวณเอง]"
        return '\n'.join(lines_out) + '\n' + note

    def _tool_item_capability(self, item_code: str = None, week: str = None) -> str:
        """item ทอได้ที่กลุ่มไหนบ้าง (จาก Table_Item) + สถานะว่างของแต่ละกลุ่ม"""
        if not item_code:
            return "ต้องระบุ item_code"
        cached, ready = get_data_cache().get('query_item_capability')
        if not ready or not cached:
            return "ข้อมูล Table_Item ยังไม่พร้อม กรุณาลองใหม่หลัง cache refresh รอบถัดไป"
        csv_text = cached.get('data', '')

        ic = item_code.strip().upper()
        exact, prefix, substr = [], [], []
        for line in csv_text.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 4:
                continue
            item = cols[0].strip()
            row = (item, cols[1].strip(), cols[2].strip(), cols[3].strip())
            iu = item.upper()
            if iu == ic:
                exact.append(row)
            elif iu.startswith(ic):
                prefix.append(row)
            elif ic in iu:
                substr.append(row)
        matched = exact or prefix or substr
        if not matched:
            return f"ไม่พบ item {item_code} ใน Table_Item"
        actual = matched[0][0]
        cap_rows = [r for r in matched if r[0] == actual]

        yw_target = self._parse_week_arg(week) if week else None
        if week and not yw_target:
            return f"ไม่เข้าใจสัปดาห์ '{week}'"

        # Ava ต่อ (group, yw) จาก Table_MC
        cap_cached, cap_ready = get_data_cache().get('query_cap_ava')
        mc_data = cap_cached.get('data', {}) if (cap_ready and cap_cached) else {}
        mc_csv = mc_data.get('mc', '') if isinstance(mc_data, dict) else ''
        avas: dict = {}
        for line in mc_csv.splitlines()[1:] if mc_csv else []:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            try:
                ava = float(cols[6].strip() or 0)
            except ValueError:
                continue
            key = (cols[1].strip().lower(), cols[0].strip())
            avas[key] = avas.get(key, 0.0) + ava

        min_yw = _min_plannable_yw()

        def _ava_for(group: str, yw: str) -> float:
            gl = group.lower()
            return sum(a for (g2, y2), a in avas.items() if y2 == yw and _group_matches(g2, [gl]))

        def _next_free(group: str):
            gl = group.lower()
            per_yw: dict = {}
            for (g2, y2), a in avas.items():
                if y2 >= min_yw and _group_matches(g2, [gl]):
                    per_yw[y2] = per_yw.get(y2, 0.0) + a
            for y2 in sorted(per_yw):
                if per_yw[y2] > 0:
                    return y2, per_yw[y2]
            return None, 0.0

        # กลุ่มที่ item มีแผนปัจจุบัน (จาก BookingMaster) — ไว้เทียบ "ทอได้" vs "มีแผนจริง"
        planned_groups = set()
        item_cached, item_ready = get_data_cache().get('query_item')
        if item_ready and item_cached:
            for line in (item_cached.get('data', '') or '').splitlines()[1:]:
                cols = line.split(',')
                if len(cols) >= 2 and cols[0].strip().upper() == actual.upper():
                    planned_groups.add(cols[1].strip().lower())

        def _is_planned(grp: str) -> str:
            gl = grp.lower()
            return 'yes' if any(
                _group_matches(pg, [gl]) or _group_matches(gl, [pg]) for pg in planned_groups
            ) else ''

        if yw_target:
            lines_out = [f"Group,SubGroup,Cap,Planned,Ava_{yw_target}"]
            for (_item, grp, sub, cap) in cap_rows[:20]:
                lines_out.append(f"{grp},{sub},{cap},{_is_planned(grp)},{_ava_for(grp, yw_target):g}")
        else:
            lines_out = ["Group,SubGroup,Cap,Planned,Next_Free_YW,Next_Free_Ava"]
            for (_item, grp, sub, cap) in cap_rows[:20]:
                nf_yw, nf_ava = _next_free(grp)
                lines_out.append(f"{grp},{sub},{cap},{_is_planned(grp)},{nf_yw or 'ไม่มีว่าง'},{nf_ava:g}")

        note = (
            f"[item {actual} ทอได้ {len(cap_rows)} กลุ่ม (จาก Table_Item); "
            f"Planned=yes คือกลุ่มที่มีแผนปัจจุบันใน BookingMaster; "
            f"Next_Free_YW นับจาก YW {min_yw} (กฎ current+2); Cap = ค่าจาก Table_Item]"
        )
        return '\n'.join(lines_out) + '\n' + note

    def _tool_get_sales_summary(self, sales_name: str = None, week: str = None) -> str:
        cached, ready = get_data_cache().get('query_sales')
        if not ready or not cached:
            return "ข้อมูล Sales ยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        sales_map = cached.get('data', {}) or {}
        entries = {k: v for k, v in sales_map.items() if k != '__system__'}
        if not entries:
            return "ไม่มีข้อมูล Sales ในระบบ"

        yw_filter = None
        if week:
            yws = _week_keyword_to_yw(week)
            if not yws and re.match(r'^\d{6}$', week.strip()):
                yws = [week.strip()]
            if not yws:
                return f"ไม่เข้าใจช่วงเวลา '{week}' — ส่งเป็นเลข YYYYWW หรือคำพูดเช่น 'เดือนนี้', 'สัปดาห์หน้า'"
            yw_filter = sorted(set(yws))

        # cache รุ่นเก่าอาจยังไม่มี kg_yw (ช่วงเปลี่ยนผ่านก่อน refresh รอบถัดไป)
        has_yw_data = any('kg_yw' in v for v in entries.values())

        def scoped_kg(v: dict) -> float:
            if yw_filter and 'kg_yw' in v:
                return sum((v.get('kg_yw') or {}).get(yw, 0.0) for yw in yw_filter)
            return float(v.get('kg', 0.0))

        if yw_filter:
            period = f"YW {yw_filter[0]}–{yw_filter[-1]}" if len(yw_filter) > 1 else f"YW {yw_filter[0]}"
        else:
            period = "ทุกสัปดาห์ในระบบ"
        warn = "" if (not yw_filter or has_yw_data) else "\n[หมายเหตุ: ข้อมูลรายสัปดาห์ยังไม่พร้อม แสดงยอดรวมทั้งหมดแทน]"

        if sales_name:
            q = sales_name.strip().lower()
            matches = [k for k in entries if k == q] or sorted(k for k in entries if q in k or k in q)
            if not matches:
                names = ', '.join(sorted(v['name'] for v in entries.values())[:15])
                return f"ไม่พบ sales '{sales_name}' — sales ที่มีในระบบ เช่น: {names}"
            lines = ["Sales,Items_All,KG"]
            for k in matches[:5]:
                v = entries[k]
                lines.append(f"{v['name']},{v['item_count']},{scoped_kg(v):.1f}")
            note = f"[ช่วงข้อมูล: {period}; Items_All = จำนวน item รวมทุกสัปดาห์]{warn}"
            return '\n'.join(lines) + '\n' + note

        ranked = sorted(entries.values(), key=lambda v: -scoped_kg(v))
        lines = ["Sales,Items_All,KG"]
        for v in ranked[:20]:
            lines.append(f"{v['name']},{v['item_count']},{scoped_kg(v):.1f}")
        total_kg = sum(scoped_kg(v) for v in entries.values())
        lines.append(f"(รวม {len(entries)} sales),,{total_kg:.1f}")
        note = f"[ช่วงข้อมูล: {period}; เรียงตาม KG แสดงสูงสุด 20 ราย; Items_All = จำนวน item รวมทุกสัปดาห์]{warn}"
        return '\n'.join(lines) + '\n' + note

    def _tool_group_utilization(self, week: str = None, weeks_ahead=None, sort_by: str = None) -> str:
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return "ข้อมูลกำลังการผลิตยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        data = cached.get('data', {})
        csv_text = data.get('mc', '') if isinstance(data, dict) else ''
        if not csv_text:
            return "ไม่มีข้อมูล Machine Capacity"

        if week:
            yws = _week_keyword_to_yw(week)
            if not yws and re.match(r'^\d{6}$', week.strip()):
                yws = [week.strip()]
            if not yws:
                return f"ไม่เข้าใจช่วงเวลา '{week}'"
            yw_list = sorted(set(yws))
        else:
            try:
                n = int(weeks_ahead) if weeks_ahead is not None else 4
            except (TypeError, ValueError):
                n = 4
            n = max(1, min(n, 12))
            now = datetime.now()
            yw_list = [_yw_from_date(now + timedelta(weeks=k)) for k in range(n)]
        yw_set = set(yw_list)

        kg_lookup = self._kg_ava_lookup(data)

        # รวมต่อ (group, yw) ก่อน เพื่อหา worst week แล้วค่อยรวมเป็นต่อ group
        per_gyw: dict = {}
        for line in csv_text.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            yw = cols[0].strip()
            if yw not in yw_set:
                continue
            gname = cols[1].strip()
            try:
                total = float(cols[3].strip() or 0)
                used = float(cols[4].strip() or 0) + float(cols[5].strip() or 0)
                ava = float(cols[6].strip() or 0)
            except ValueError:
                continue
            c = per_gyw.setdefault((gname, yw), [0.0, 0.0, 0.0])
            c[0] += total
            c[1] += used
            c[2] += ava

        if not per_gyw:
            return f"ไม่พบข้อมูล capacity ในช่วง YW {yw_list[0]}–{yw_list[-1]}"

        per_group: dict = {}
        for (gname, yw), (total, used, ava) in per_gyw.items():
            g = per_group.setdefault(gname, {'total': 0.0, 'used': 0.0, 'ava': 0.0, 'worst': None})
            g['total'] += total
            g['used'] += used
            g['ava'] += ava
            if g['worst'] is None or ava < g['worst'][1]:
                g['worst'] = (yw, ava)

        def _used_pct(g):
            return (g['used'] / g['total'] * 100) if g['total'] > 0 else 0.0

        reverse_key = (lambda kv: -kv[1]['ava']) if (sort_by or '').lower() in ('available', 'ava', 'ว่าง') \
            else (lambda kv: -_used_pct(kv[1]))
        ranked = sorted(per_group.items(), key=reverse_key)

        lines = ["Group,Total,Used,Ava,Used_Pct,KG_Ava,Worst_YW,Worst_Ava"]
        for gname, g in ranked[:30]:
            kg = sum(kg_lookup.get((yw, gname.lower()), 0.0) for yw in yw_list)
            kg_s = f"{kg:.2f}" if kg > 0 else ''
            worst_yw, worst_ava = g['worst']
            lines.append(
                f"{gname},{g['total']:g},{g['used']:g},{g['ava']:g},"
                f"{_used_pct(g):.1f}%,{kg_s},{worst_yw},{worst_ava:g}"
            )
        note = (
            f"[ช่วงข้อมูล: YW {yw_list[0]}–{yw_list[-1]} ({len(yw_list)} สัปดาห์) "
            f"เรียงตาม{'เครื่องว่างมาก' if (sort_by or '').lower() in ('available', 'ava', 'ว่าง') else ' Used% มาก'}; "
            f"ใช้ Used_Pct จากตารางนี้เท่านั้น ห้ามคำนวณเอง; Used_Pct > 100% = จองเกิน capacity]"
        )
        return '\n'.join(lines) + '\n' + note

    def _tool_analyze_plan_impact(self, item_code: str = None, week: str = None, group: str = None) -> str:
        """วิเคราะห์ผลกระทบถ้าถอด/ย้ายแผนของ item: capacity กลุ่มเปลี่ยนเท่าไร และ item ใดใช้กลุ่ม+สัปดาห์ร่วมกัน"""
        if not item_code:
            return "ต้องระบุ item_code ที่จะวิเคราะห์ผลกระทบ"
        item_cached, item_ready = get_data_cache().get('query_item')
        if not item_ready or not item_cached:
            return "ข้อมูล Item Plan ยังไม่พร้อม กรุณาลองใหม่อีกครั้ง"
        item_csv = item_cached.get('data', '')
        cap_cached, cap_ready = get_data_cache().get('query_cap_ava')
        cap_data = cap_cached.get('data', {}) if (cap_ready and cap_cached) else {}
        mc_csv = cap_data.get('mc', '') if isinstance(cap_data, dict) else ''

        yw_filter = None
        if week:
            yws = _week_keyword_to_yw(week)
            if not yws and re.match(r'^\d{6}$', week.strip()):
                yws = [week.strip()]
            yw_filter = set(yws) if yws else None

        # หาแผนของ item (exact > prefix > substring แบบเดียวกับ get_item_plan)
        # พร้อมเก็บทุกแถวไว้หา item ที่ใช้กลุ่ม+สัปดาห์ร่วมกัน
        ic = item_code.strip().upper()
        exact, prefix, substr = [], [], []
        all_rows = []  # (item, group, kp, yw)
        for line in item_csv.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 4:
                continue
            item, grp, kp_s, yw = cols[0].strip(), cols[1].strip(), cols[2].strip(), cols[3].strip()
            try:
                kp = float(kp_s) if kp_s else 0.0
            except ValueError:
                kp = 0.0
            row = (item, grp, kp, yw)
            all_rows.append(row)
            iu = item.upper()
            if iu == ic:
                exact.append(row)
            elif iu.startswith(ic):
                prefix.append(row)
            elif ic in iu:
                substr.append(row)

        matched = exact or prefix or substr
        if not matched:
            return f"ไม่พบ item {item_code} ใน Item Plan"
        actual_code = matched[0][0]
        target_rows = [r for r in matched if r[0] == actual_code]
        if group:
            target_rows = [r for r in target_rows if _group_matches(r[1].lower(), [group.lower()])]
        if yw_filter:
            target_rows = [r for r in target_rows if r[3] in yw_filter]
        if not target_rows:
            return f"item {actual_code} ไม่มีแผนตามเงื่อนไข (week={week}, group={group})"

        # capacity รวมต่อ (yw, group) จาก Table_MC
        cap: dict = {}
        for line in mc_csv.splitlines()[1:] if mc_csv else []:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            key = (cols[0].strip(), cols[1].strip().lower())
            c = cap.setdefault(key, [0.0, 0.0, 0.0])  # Total, Used, Ava
            try:
                c[0] += float(cols[3].strip() or 0)
                c[1] += float(cols[4].strip() or 0) + float(cols[5].strip() or 0)
                c[2] += float(cols[6].strip() or 0)
            except ValueError:
                pass

        _MAX_PLAN_ROWS = 8
        out = [f"=== วิเคราะห์ผลกระทบแผนของ {actual_code} ==="]
        for (item, grp, kp, yw) in target_rows[:_MAX_PLAN_ROWS]:
            grp_l = grp.lower()
            others = sorted(
                ((i, k) for (i, g2, k, y2) in all_rows if y2 == yw and g2.lower() == grp_l and i != actual_code),
                key=lambda x: -x[1],
            )
            group_kp = kp + sum(k for _, k in others)
            out.append(f"--- ถ้าถอดแผน {actual_code} YW {yw} กลุ่ม {grp} ({kp:g} kg) ---")
            c = cap.get((yw, grp_l))
            if c:
                total, used, ava = c
                status = 'เต็ม/เกิน' if ava <= 0 else 'ยังมีว่าง'
                out.append(f"Capacity กลุ่ม {grp} YW {yw}: Total={total:g}, Used={used:g}, Ava={ava:g} ({status})")
                if used > 0 and group_kp > 0:
                    est = kp / (group_kp / used)  # ประมาณจากสัดส่วน KP_Weight ของกลุ่มในสัปดาห์นั้น
                    out.append(f"ประมาณเครื่องที่แผนนี้ใช้ ~{est:.1f} เครื่อง → ถ้าถอดออก Ava เพิ่มเป็น ~{ava + est:.1f}")
            else:
                out.append(f"(ไม่พบข้อมูล capacity ของกลุ่ม {grp} YW {yw})")
            if others:
                out.append(
                    f"item ที่ใช้กลุ่ม {grp} YW {yw} ร่วมกัน: {len(others)} รายการ "
                    f"รวม {sum(k for _, k in others):.2f} kg — ได้เครื่องว่างเพิ่มถ้าถอด / ถูกเบียดถ้าเพิ่มแผน"
                )
                out.append("  Top: " + ', '.join(f"{i} ({k:g} kg)" for i, k in others[:5]))
            else:
                out.append(f"ไม่มี item อื่นในกลุ่ม {grp} YW {yw} — ถอดแผนนี้ไม่กระทบ item อื่น")
            out.append("")
        if len(target_rows) > _MAX_PLAN_ROWS:
            out.append(f"[หมายเหตุ: item นี้มีแผน {len(target_rows)} รายการ แสดง {_MAX_PLAN_ROWS} รายการแรก — ระบุ week/group เพื่อเจาะจง]")
        out.append(
            f"[หมายเหตุ: 'เครื่องที่แผนนี้ใช้' เป็นค่าประมาณจากสัดส่วน KP_Weight; "
            f"ถ้าจะย้ายแผนไป week อื่น week เร็วที่สุดที่วางแผนได้ = YW {_min_plannable_yw()}]"
        )
        return '\n'.join(out)

    _EXCEL_EXPORT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'excel_exports')
    _EXCEL_MAX_ROWS = 5000

    # คอลัมน์ที่ต้องคงเป็น string ตอนแปลง CSV → Excel (รหัส/สัปดาห์ ไม่ใช่ตัวเลขเชิงปริมาณ)
    _EXPORT_TEXT_COLS = {'yw', 'item', 'group', 'guage', 'mc_group', 'groups'}

    @classmethod
    def _csv_to_table(cls, text: str):
        """แปลงผลลัพธ์ CSV จาก data tool เป็น (columns, rows) สำหรับ export
        ข้ามบรรทัด note/footer ที่ขึ้นต้นด้วย '[' — คืน ([], []) ถ้าไม่ใช่ตาราง"""
        lines = [l for l in (text or '').splitlines() if l.strip() and not l.strip().startswith('[')]
        if len(lines) < 2 or ',' not in lines[0]:
            return [], []
        columns = [c.strip() for c in lines[0].split(',')]
        text_idx = {i for i, c in enumerate(columns) if c.lower() in cls._EXPORT_TEXT_COLS}
        rows = []
        for line in lines[1:]:
            cells = [c.strip() for c in line.split(',')]
            cells = cells[:len(columns)] + [''] * (len(columns) - len(cells))
            row = []
            for i, v in enumerate(cells):
                if i not in text_idx and v:
                    try:
                        v = float(v)
                    except ValueError:
                        pass
                row.append(v)
            rows.append(row)
        return columns, rows

    def _run_data_tool_for_export(self, source_tool: str, args: dict):
        """เรียก data tool ฝั่ง server เพื่อเอา CSV มา export ตรงๆ — คืน None ถ้าไม่รู้จัก tool"""
        if source_tool in ('get_item_plan', 'get_knit_plan'):
            return self._tool_get_item_plan(
                item_code=args.get('item_code'), group=args.get('group'), week=args.get('week'))
        if source_tool == 'get_machine_capacity':
            return self._tool_get_machine_capacity(
                group=args.get('group'), week=args.get('week'), gauge=args.get('gauge'))
        if source_tool == 'get_booking':
            return self._tool_get_booking(group=args.get('group'), week=args.get('week'))
        return None

    def _tool_export_excel(self, raw_arguments: str) -> str:
        """สร้างไฟล์ .xlsx ใน excel_exports/ แล้วคืนลิงก์ดาวน์โหลดให้โมเดลใส่ในคำตอบ"""
        try:
            args = json.loads(raw_arguments)
        except Exception:
            # arguments มัก truncate เมื่อโมเดลพยายามส่ง rows เองจำนวนมาก — ชี้ทางให้ใช้ source_tool
            return ("สร้างไฟล์ Excel ไม่สำเร็จ: arguments ไม่สมบูรณ์ — "
                    "ให้เรียกใหม่ด้วย source_tool + filter แทนการส่ง rows เอง")

        columns = [str(c) for c in (args.get("columns") or [])]
        rows = [r for r in (args.get("rows") or []) if isinstance(r, list)]

        source_tool = args.get("source_tool")
        if source_tool and not rows:
            csv_text = self._run_data_tool_for_export(str(source_tool), args)
            if csv_text is None:
                return f"สร้างไฟล์ Excel ไม่สำเร็จ: ไม่รู้จัก source_tool '{source_tool}'"
            columns, rows = self._csv_to_table(csv_text)
            if not rows:
                first_line = (csv_text or '').splitlines()[0] if csv_text else 'ไม่มีข้อมูล'
                return f"สร้างไฟล์ Excel ไม่สำเร็จ: {first_line}"

        if not columns or not rows:
            return "สร้างไฟล์ Excel ไม่สำเร็จ: ต้องระบุ source_tool + filter หรือส่ง columns และ rows ที่มีข้อมูล"
        truncated = len(rows) > self._EXCEL_MAX_ROWS
        rows = rows[:self._EXCEL_MAX_ROWS]

        slug = re.sub(r'[^A-Za-z0-9_-]+', '', str(args.get("filename") or '')) or 'report'
        # timestamp เวลาไทย — container รัน UTC ถ้าใช้ datetime.now() เฉยๆ ชื่อไฟล์จะเพี้ยน 7 ชม.
        now_bkk = datetime.now(timezone(timedelta(hours=7)))
        fname = f"{slug}_{now_bkk.strftime('%Y%m%d_%H%M%S')}.xlsx"

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = 'Report'
            title = str(args.get("title") or '')
            row0 = 1
            if title:
                ws.cell(1, 1, title).font = Font(bold=True, size=13)
                row0 = 3
            for ci, col in enumerate(columns, 1):
                ws.cell(row0, ci, col).font = Font(bold=True)
            for ri, row in enumerate(rows, row0 + 1):
                for ci, v in enumerate(row[:len(columns)], 1):
                    ws.cell(ri, ci, v if isinstance(v, (int, float)) else ('' if v is None else str(v)))
            # ปรับความกว้างคอลัมน์ตามเนื้อหา (cap ที่ 40)
            for ci, col in enumerate(columns, 1):
                width = len(str(col))
                for row in rows:
                    if len(row) >= ci:
                        width = max(width, len(str(row[ci - 1])))
                ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 40)

            os.makedirs(self._EXCEL_EXPORT_DIR, exist_ok=True)
            wb.save(os.path.join(self._EXCEL_EXPORT_DIR, fname))
        except Exception as e:
            logger.error(f"export_excel failed: {e}")
            return "สร้างไฟล์ Excel ไม่สำเร็จ: เกิดข้อผิดพลาดภายใน กรุณาลองใหม่อีกครั้ง"

        logger.info(f"export_excel: {fname} | {len(columns)} cols | {len(rows)} rows")
        trunc_note = f" (ตัดเหลือ {self._EXCEL_MAX_ROWS} แถวแรก)" if truncated else ""
        return (
            f"สร้างไฟล์ Excel เรียบร้อย ({len(rows)} แถว{trunc_note}) — "
            f"ตอบ user พร้อมแนบลิงก์บรรทัดนี้เป๊ะๆ ห้ามแก้ URL:\n"
            f"[📥 ดาวน์โหลด {fname}](/download/excel/{fname})"
        )

    def _execute_tool_call(self, tool_call) -> str:
        name = tool_call.function.name
        try:
            args = json.loads(tool_call.function.arguments)
        except Exception:
            args = {}
        logger.info(f"Tool call: {name}({args})")

        if name == "get_item_plan":
            return self._tool_get_item_plan(
                item_code=args.get("item_code"),
                group=args.get("group"),
                week=args.get("week"),
            )
        elif name == "get_machine_capacity":
            return self._tool_get_machine_capacity(
                group=args.get("group"),
                week=args.get("week"),
                gauge=args.get("gauge"),
            )
        elif name == "get_booking":
            return self._tool_get_booking(
                group=args.get("group"),
                week=args.get("week"),
            )
        elif name == "get_knit_plan":
            return self._tool_get_knit_plan(
                week=args.get("week"),
                group=args.get("group"),
                item_code=args.get("item_code"),
            )
        elif name == "suggest_week":
            return self._tool_suggest_week(
                group=args.get("group"),
                gauge=args.get("gauge"),
                machines_needed=args.get("machines_needed"),
            )
        elif name == "compare_weeks":
            return self._tool_compare_weeks(
                week_a=args.get("week_a"),
                week_b=args.get("week_b"),
                group=args.get("group"),
            )
        elif name == "item_capability":
            return self._tool_item_capability(
                item_code=args.get("item_code"),
                week=args.get("week"),
            )
        elif name == "get_sales_summary":
            return self._tool_get_sales_summary(
                sales_name=args.get("sales_name"),
                week=args.get("week"),
            )
        elif name == "group_utilization":
            return self._tool_group_utilization(
                week=args.get("week"),
                weeks_ahead=args.get("weeks_ahead"),
                sort_by=args.get("sort_by"),
            )
        elif name == "analyze_plan_impact":
            return self._tool_analyze_plan_impact(
                item_code=args.get("item_code"),
                week=args.get("week"),
                group=args.get("group"),
            )
        elif name == "export_excel":
            return self._tool_export_excel(tool_call.function.arguments)
        return f"ไม่รู้จัก tool: {name}"

    def _build_table_spec(self, raw_arguments: str):
        """แปลง arguments จาก render_table เป็น table spec ที่ frontend ใช้แสดง.
        คืน (spec | None, ข้อความผลลัพธ์สำหรับป้อนกลับให้โมเดล)."""
        try:
            args = json.loads(raw_arguments)
        except Exception:
            return None, "สร้างตารางไม่สำเร็จ: ข้อมูลไม่ถูกต้อง"

        columns = [str(c) for c in (args.get("columns") or [])]
        raw_rows = args.get("rows") or []
        rows = []
        for row in raw_rows:
            if isinstance(row, list):
                rows.append([v if isinstance(v, (int, float)) else (str(v) if v is not None else '') for v in row])

        if not columns or not rows:
            return None, "สร้างตารางไม่สำเร็จ: ต้องมี columns และ rows ที่มีข้อมูล"

        spec = {
            "title": str(args.get("title") or ""),
            "columns": columns,
            "rows": rows,
        }
        logger.info(f"render_table: {len(columns)} cols | {len(rows)} rows")
        return spec, "สร้างตารางเรียบร้อยแล้ว แสดงให้ผู้ใช้ทางหน้าจอแล้ว — ให้เขียนคำอธิบายสั้นๆ ประกอบตาราง ห้ามลิสต์ตัวเลขทั้งหมดซ้ำ"

    def _build_chart_spec(self, raw_arguments: str):
        """แปลง arguments จาก render_chart เป็น chart spec ที่ frontend ใช้วาด (Chart.js).
        คืน (spec | None, ข้อความผลลัพธ์สำหรับป้อนกลับให้โมเดล)."""
        try:
            args = json.loads(raw_arguments)
        except Exception:
            return None, "สร้างกราฟไม่สำเร็จ: ข้อมูลกราฟไม่ถูกต้อง"

        chart_type = (args.get("chart_type") or "bar").lower()
        if chart_type not in ("bar", "line", "pie", "doughnut"):
            chart_type = "bar"

        labels = [str(x) for x in (args.get("labels") or [])]
        raw_datasets = args.get("datasets") or []
        datasets = []
        for ds in raw_datasets:
            if not isinstance(ds, dict):
                continue
            values = []
            for v in (ds.get("data") or []):
                try:
                    values.append(round(float(v), 4))
                except (TypeError, ValueError):
                    values.append(0)
            if values:
                datasets.append({"label": str(ds.get("label") or "ข้อมูล"), "data": values})

        if not labels or not datasets:
            return None, "สร้างกราฟไม่สำเร็จ: ต้องมี labels และ datasets ที่มีค่าตัวเลข"

        spec = {
            "type": chart_type,
            "title": str(args.get("title") or ""),
            "labels": labels,
            "datasets": datasets,
            "y_label": str(args.get("y_label") or ""),
        }
        logger.info(f"render_chart: {chart_type} | {len(labels)} labels | {len(datasets)} datasets")
        return spec, "สร้างกราฟเรียบร้อยแล้ว แสดงให้ผู้ใช้ทางหน้าจอแล้ว — ให้เขียนคำอธิบายสั้นๆ ประกอบกราฟ ห้ามลิสต์ตัวเลขทั้งหมดซ้ำ"

    _MACHINE_INFO_REPLY = (
        "ตอนนี้น้อง I-SAVE Chatbot ยังไม่รู้เลยว่าพี่ๆ ต้องการดูข้อมูลเครื่องจักรส่วนไหนค่ะ เช่น\n"
        "• เครื่องว่าง (Ava) ตามสัปดาห์\n"
        "• กำลังการผลิตรวมของแต่ละกลุ่มเครื่อง\n"
        "• การจองเครื่องของแต่ละกลุ่ม / สัปดาห์\n"
        "• หรือข้อมูลของกลุ่มเครื่องใดกลุ่มหนึ่ง (เช่น SKP, SKPLE, SKPTA)\n"
        "รบกวนพิมพ์เพิ่มให้หน่อยนะคะว่าอยากดู\n"
        "1) กลุ่มเครื่องอะไร (ถ้ามี) และ\n"
        "2) สัปดาห์ไหน (เช่น 202624 หรือ ช่วงสัปดาห์ 202624–202626)\n"
        "แล้วน้องจะดึงข้อมูลจากระบบ I-SAVE ให้ทันทีค่ะ"
    )

    _CAPACITY_INFO_REPLY = (
        "ขออภัยค่ะ ข้อมูล Capacity ยังไม่พร้อมในขณะนี้\n"
        "กรุณารอสักครู่แล้วลองใหม่อีกครั้ง หรือติดต่อผู้ดูแลระบบหากปัญหายังคงอยู่ค่ะ"
    )

    def _get_machine_capacity_suggestions(self) -> list:
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return ['เครื่องว่าง SKP สัปดาห์นี้', 'กำลังการผลิตรวม', 'การจองเครื่อง', 'ดูแผนทอ']
        data = cached.get('data', {})
        mc_csv = data.get('mc', '') if isinstance(data, dict) else ''
        if not mc_csv:
            return ['เครื่องว่าง SKP สัปดาห์นี้', 'กำลังการผลิตรวม', 'การจองเครื่อง', 'ดูแผนทอ']

        min_yw = _min_plannable_yw()
        rows = []
        for line in mc_csv.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) >= 7:
                try:
                    rows.append({
                        'yw': cols[0].strip(),
                        'group': cols[1].strip(),
                        'total': float(cols[3]) if cols[3].strip() else 0,
                        'ava': float(cols[6]) if cols[6].strip() else 0,
                    })
                except ValueError:
                    pass

        snapshot_rows = [r for r in rows if r['yw'] == min_yw]
        if not snapshot_rows:
            future = [r for r in rows if r['yw'] >= min_yw]
            if future:
                first_yw = min(r['yw'] for r in future)
                snapshot_rows = [r for r in rows if r['yw'] == first_yw]
        if not snapshot_rows:
            return ['เครื่องว่าง SKP สัปดาห์นี้', 'กำลังการผลิตรวม', 'การจองเครื่อง', 'ดูแผนทอ']

        snapshot_yw = snapshot_rows[0]['yw']
        try:
            week_num = int(snapshot_yw[4:]) if len(snapshot_yw) == 6 else int(snapshot_yw)
        except ValueError:
            week_num = 0

        group_ava: dict = {}
        for r in snapshot_rows:
            group_ava[r['group']] = group_ava.get(r['group'], 0) + r['ava']

        sorted_groups = sorted(group_ava, key=lambda g: group_ava[g], reverse=True)
        suggestions = []
        for g in sorted_groups[:2]:
            suggestions.append(f"Capacity กลุ่ม {g} สัปดาห์ที่ {week_num}")
        suggestions.append(f"เครื่องว่างทุกกลุ่ม สัปดาห์ที่ {week_num}")
        suggestions.append("การจองเครื่องทั้งหมด")
        return suggestions[:4]

    def _build_machine_capacity_overview(self) -> str:
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return self._MACHINE_INFO_REPLY

        data = cached.get('data', {})
        mc_csv = data.get('mc', '') if isinstance(data, dict) else ''
        if not mc_csv:
            return self._MACHINE_INFO_REPLY

        lines = mc_csv.splitlines()
        # columns: YW,Group,Guage,Total,Used_N,Used_F,Ava

        rows = []
        for line in lines[1:]:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            try:
                rows.append({
                    'yw': cols[0].strip(),
                    'group': cols[1].strip(),
                    'guage': cols[2].strip(),
                    'total': float(cols[3]) if cols[3].strip() else 0,
                    'used_n': float(cols[4]) if cols[4].strip() else 0,
                    'used_f': float(cols[5]) if cols[5].strip() else 0,
                    'ava': float(cols[6]) if cols[6].strip() else 0,
                })
            except ValueError:
                continue

        if not rows:
            return self._MACHINE_INFO_REPLY

        min_yw = _min_plannable_yw()

        # Use nearest plannable week as snapshot
        snapshot_rows = [r for r in rows if r['yw'] == min_yw]
        if not snapshot_rows:
            future_rows = [r for r in rows if r['yw'] >= min_yw]
            if future_rows:
                first_yw = min(r['yw'] for r in future_rows)
                snapshot_rows = [r for r in rows if r['yw'] == first_yw]
        if not snapshot_rows:
            snapshot_rows = rows

        snapshot_yw = snapshot_rows[0]['yw'] if snapshot_rows else min_yw
        try:
            week_num = int(snapshot_yw[4:]) if len(snapshot_yw) == 6 else int(snapshot_yw)
        except ValueError:
            week_num = 0

        # Distinct main groups and sub-groups across all data
        all_groups = set(r['group'] for r in rows if r['group'])
        all_subgroups = set((r['group'], r['guage']) for r in rows if r['group'])

        # Aggregate by group for snapshot week
        group_total: dict = {}
        group_ava: dict = {}
        for r in snapshot_rows:
            g = r['group']
            group_total[g] = group_total.get(g, 0) + r['total']
            group_ava[g] = group_ava.get(g, 0) + r['ava']

        total_machines = sum(group_total.values())
        total_ava = sum(group_ava.values())

        group_util = {
            g: (group_total[g] - group_ava.get(g, 0)) / group_total[g]
            for g in group_total if group_total[g] > 0
        }

        max_machines_group = max(group_total, key=lambda g: group_total[g]) if group_total else '-'
        tightest_group = max(group_util, key=lambda g: group_util[g]) if group_util else '-'

        # KP Weight total from item plan for snapshot week
        total_kg = 0.0
        item_cached, item_ready = get_data_cache().get('query_item')
        if item_ready and item_cached:
            item_data = item_cached.get('data', '')
            for item_line in item_data.splitlines()[1:]:
                item_cols = item_line.split(',')
                if len(item_cols) >= 4 and item_cols[3].strip() == snapshot_yw:
                    try:
                        total_kg += float(item_cols[2].strip())
                    except ValueError:
                        pass

        # KG_Ava รวมของ snapshot week
        total_kg_ava = 0.0
        kg_ava_csv = data.get('kg_ava', '') if isinstance(data, dict) else ''
        if kg_ava_csv:
            for ka_line in kg_ava_csv.splitlines()[1:]:
                ka_cols = ka_line.split(',')
                if len(ka_cols) >= 3 and ka_cols[0].strip() == snapshot_yw:
                    try:
                        total_kg_ava += float(ka_cols[2].strip())
                    except ValueError:
                        pass

        n_main = len(all_groups)
        n_sub = len(all_subgroups)
        tightest_util_pct = round(group_util.get(tightest_group, 0) * 100)
        tightest_ava = int(group_ava.get(tightest_group, 0))

        out_lines = [
            f"ภาพรวม Machine Capacity สัปดาห์ที่ {week_num}",
            f"มีเครื่องทั้งหมด **{int(total_machines)} เครื่อง** แบ่งเป็น **{n_main} กลุ่มเครื่องหลัก** และ **{n_sub} กลุ่มย่อย** ค่ะ",
        ]
        if total_kg > 0:
            out_lines.append(f"KP Weight รวมสัปดาห์นี้อยู่ที่ประมาณ **{total_kg:,.0f} kg**")
        out_lines.append(
            f"กลุ่มที่มีเครื่องมากที่สุดคือ **{max_machines_group}** ({int(group_total.get(max_machines_group, 0))} เครื่อง)"
        )
        if tightest_group and tightest_group != '-':
            out_lines.append(
                f"กลุ่มที่ Capacity ค่อนข้างตึงคือ **{tightest_group}** "
                f"(ใช้ไปแล้ว {tightest_util_pct}% เครื่องว่าง {tightest_ava} เครื่อง)"
            )
        ava_line = f"เครื่องว่างรวม **{int(total_ava)} เครื่อง** จาก {int(total_machines)} เครื่องทั้งหมด"
        if total_kg_ava > 0 and total_ava > 0:
            ava_line += f" (KG_Ava รวม **{total_kg_ava:,.0f} kg**)"
        out_lines.append(ava_line)
        out_lines.append("ถ้าต้องการดูรายละเอียดเพิ่มเติม สามารถเลือกกลุ่มเครื่อง หรือระบุสัปดาห์ที่ต้องการได้เลยค่ะ")

        return '\n'.join(out_lines)


    def _parse_mc_snapshot(self):
        """Return (snapshot_rows, snapshot_yw, week_num, group_total, group_used, group_ava) or None if no data."""
        cached, ready = get_data_cache().get('query_cap_ava')
        if not ready or not cached:
            return None
        data = cached.get('data', {})
        mc_csv = data.get('mc', '') if isinstance(data, dict) else ''
        if not mc_csv:
            return None

        min_yw = _min_plannable_yw()
        rows = []
        for line in mc_csv.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 7:
                continue
            try:
                rows.append({
                    'yw': cols[0].strip(), 'group': cols[1].strip(),
                    'total': float(cols[3]) if cols[3].strip() else 0,
                    'used_n': float(cols[4]) if cols[4].strip() else 0,
                    'used_f': float(cols[5]) if cols[5].strip() else 0,
                    'ava': float(cols[6]) if cols[6].strip() else 0,
                })
            except ValueError:
                continue
        if not rows:
            return None

        snap = [r for r in rows if r['yw'] == min_yw]
        if not snap:
            future = [r for r in rows if r['yw'] >= min_yw]
            if future:
                first_yw = min(r['yw'] for r in future)
                snap = [r for r in rows if r['yw'] == first_yw]
        if not snap:
            snap = rows

        snapshot_yw = snap[0]['yw']
        try:
            week_num = int(snapshot_yw[4:]) if len(snapshot_yw) == 6 else int(snapshot_yw)
        except ValueError:
            week_num = 0

        group_total: dict = {}
        group_used: dict = {}
        group_ava: dict = {}
        for r in snap:
            g = r['group']
            group_total[g] = group_total.get(g, 0) + r['total']
            group_used[g] = group_used.get(g, 0) + r['used_n'] + r['used_f']
            group_ava[g] = group_ava.get(g, 0) + r['ava']

        return snap, snapshot_yw, week_num, group_total, group_used, group_ava

    def _build_capacity_overview(self) -> str:
        parsed = self._parse_mc_snapshot()
        if not parsed:
            return self._CAPACITY_INFO_REPLY

        _, snapshot_yw, week_num, group_total, group_used, group_ava = parsed

        total_machines = sum(group_total.values())
        total_used = sum(group_used.values())
        total_ava = sum(group_ava.values())
        util_pct = (total_used / total_machines * 100) if total_machines > 0 else 0
        n_groups = len(group_total)
        min_yw = _min_plannable_yw()

        # KP Weight: sum across ALL plannable weeks = total planned load
        total_kp_all = 0.0
        snapshot_kp = 0.0
        item_cached, item_ready = get_data_cache().get('query_item')
        if item_ready and item_cached:
            item_data = item_cached.get('data', '')
            for item_line in item_data.splitlines()[1:]:
                cols = item_line.split(',')
                if len(cols) < 4:
                    continue
                yw_col = cols[3].strip()
                kp_raw = cols[2].strip()
                if yw_col >= min_yw and kp_raw:
                    try:
                        kp = float(kp_raw)
                        total_kp_all += kp
                        if yw_col == snapshot_yw:
                            snapshot_kp += kp
                    except ValueError:
                        pass

        remaining_kp = total_kp_all - snapshot_kp

        # KG_Ava รวมของ snapshot week
        total_kg_ava = 0.0
        cap_cached, _ = get_data_cache().get('query_cap_ava')
        if cap_cached:
            kg_ava_csv = cap_cached.get('data', {}).get('kg_ava', '') if isinstance(cap_cached.get('data'), dict) else ''
            for ka_line in kg_ava_csv.splitlines()[1:]:
                ka_cols = ka_line.split(',')
                if len(ka_cols) >= 3 and ka_cols[0].strip() == snapshot_yw:
                    try:
                        total_kg_ava += float(ka_cols[2].strip())
                    except ValueError:
                        pass

        # Per-group utilization %
        group_util_pct = {
            g: (group_used[g] / group_total[g] * 100) if group_total.get(g, 0) > 0 else 0
            for g in group_total
        }
        highest_util_group = max(group_util_pct, key=lambda g: group_util_pct[g]) if group_util_pct else '-'
        most_ava_group = max(group_ava, key=lambda g: group_ava[g]) if group_ava else '-'
        critical_threshold = 85.0
        critical = {g: u for g, u in group_util_pct.items() if u > critical_threshold}
        critical_group = max(critical, key=lambda g: critical[g]) if critical else None

        out = [
            f"ตอนนี้ข้อมูล Capacity ในระบบครอบคลุมตั้งแต่สัปดาห์ที่ **{week_num}** เป็นต้นไป",
            "",
            "**ภาพรวม Capacity ทั้งหมด**",
            f"- จำนวนกลุ่มเครื่องที่มีข้อมูล Capacity : **{n_groups} กลุ่ม**",
        ]
        if total_kp_all > 0:
            out.append(f"- Capacity รวมทั้งหมด (KP Weight ทุกสัปดาห์) : **{total_kp_all:,.0f} kg**")
            out.append(f"- ใช้งานไปแล้ว (สัปดาห์ที่ {week_num}) : **{snapshot_kp:,.0f} kg**")
            out.append(f"- Capacity คงเหลือในแผน : **{remaining_kp:,.0f} kg**")
        mc_ava_line = f"- เครื่องทั้งหมด : **{int(total_machines)} เครื่อง** | ใช้งาน **{int(total_used)}** | ว่าง **{int(total_ava)}**"
        if total_kg_ava > 0 and total_ava > 0:
            mc_ava_line += f" | KG_Ava **{total_kg_ava:,.0f} kg**"
        out += [
            mc_ava_line,
            f"- % การใช้ Capacity เฉลี่ย : **{round(util_pct, 1)}%**",
            "",
            "จากภาพรวมตอนนี้",
        ]
        if highest_util_group != '-':
            out.append(f"กลุ่มที่ใช้ Capacity สูงที่สุดคือ **{highest_util_group}** (ใช้งาน {round(group_util_pct[highest_util_group], 1)}%)")
        if most_ava_group != '-':
            out.append(f"กลุ่มที่ยังมี Capacity เหลือมากที่สุดคือ **{most_ava_group}** ({int(group_ava[most_ava_group])} เครื่องว่าง)")
        if critical_group:
            out.append(f"และกลุ่มที่ควรติดตามใกล้ชิดคือ **{critical_group}** เพราะมีการใช้ Capacity เกิน {round(critical[critical_group], 1)}%")
        out.append("ถ้าต้องการดูรายละเอียดเพิ่มเติม สามารถระบุกลุ่มเครื่องหรือสัปดาห์ที่ต้องการได้เลยค่ะ")
        return '\n'.join(out)

    def _get_capacity_suggestions(self) -> list:
        parsed = self._parse_mc_snapshot()
        if not parsed:
            return ['ขอดู Capacity กลุ่ม SKP', 'เครื่องว่างรวมทุกกลุ่ม', 'การจองเครื่อง', 'ดูแผนทอ']

        _, _, week_num, group_total, group_used, _ = parsed
        group_util_pct = {
            g: (group_used[g] / group_total[g] * 100) if group_total.get(g, 0) > 0 else 0
            for g in group_total
        }
        # Top 2 by utilization — most actionable groups
        top2 = sorted(group_util_pct, key=lambda g: group_util_pct[g], reverse=True)[:2]
        suggestions = [f"Capacity กลุ่ม {g} สัปดาห์ที่ {week_num}" for g in top2]
        suggestions.append(f"เครื่องว่างทุกกลุ่ม สัปดาห์ที่ {week_num}")
        suggestions.append("KP Weight รวมทุกกลุ่ม")
        return suggestions[:4]

    def _build_item_overview(self) -> str:
        item_cached, item_ready = get_data_cache().get('query_item')
        if not item_ready or not item_cached:
            return "ข้อมูล Item Plan ยังไม่พร้อม กรุณาลองใหม่อีกครั้งค่ะ"
        item_data = item_cached.get('data', '')
        if not item_data:
            return "ไม่มีข้อมูล Item Plan ในระบบค่ะ"

        min_yw = _min_plannable_yw()
        distinct_items: set = set()
        distinct_groups: set = set()
        distinct_weeks: set = set()
        total_kp = 0.0

        for line in item_data.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 4:
                continue
            item_code = cols[0].strip()
            group = cols[1].strip()
            yw = cols[3].strip()
            if not item_code or not yw or yw < min_yw:
                continue
            try:
                kp = float(cols[2].strip()) if cols[2].strip() else 0.0
            except ValueError:
                kp = 0.0
            distinct_items.add(item_code)
            distinct_groups.add(group)
            distinct_weeks.add(yw)
            total_kp += kp

        if not distinct_items:
            return "ไม่พบข้อมูล Item Plan ในช่วงสัปดาห์ที่วางแผนได้ค่ะ"

        min_week_yw = min(distinct_weeks)
        try:
            min_week_num = int(min_week_yw[4:]) if len(min_week_yw) == 6 else int(min_week_yw)
        except ValueError:
            min_week_num = 0

        out = [
            f"ตอนนี้ระบบมีข้อมูล Item Plan ทั้งหมด **{len(distinct_items)} Item**",
            f"ครอบคลุมตั้งแต่สัปดาห์ที่ **{min_week_num}** เป็นต้นไป",
            "",
            "**ภาพรวม Item**",
            f"- จำนวน Item ที่มีแผนทอ : **{len(distinct_items)} Item**",
            f"- น้ำหนักรวมทั้งหมด : **{total_kp:,.0f} kg**",
            f"- จำนวนกลุ่มเครื่องที่เกี่ยวข้อง : **{len(distinct_groups)} กลุ่ม**",
            f"- จำนวนสัปดาห์ในแผน : **{len(distinct_weeks)} สัปดาห์**",
            "",
            'ถ้าต้องการดูละเอียดขึ้น สามารถระบุรหัส Item, กลุ่มเครื่อง หรือสัปดาห์ได้เลยค่ะ',
            'เช่น "ขอดู Item F100114/10A0" หรือ "Item ในกลุ่ม SKP สัปดาห์ที่ ' + str(min_week_num) + '"',
        ]
        return '\n'.join(out)

    def _get_item_suggestions(self) -> list:
        item_cached, item_ready = get_data_cache().get('query_item')
        if not item_ready or not item_cached:
            return ['ดู Item ในกลุ่ม SKP', 'ดูแผนทอ Item', 'ดูเครื่องว่าง', 'ข้อมูลแผนทอ']

        item_data = item_cached.get('data', '')
        min_yw = _min_plannable_yw()

        group_items: dict = {}
        min_week_yw = None
        for line in item_data.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 4:
                continue
            item_code = cols[0].strip()
            group = cols[1].strip()
            yw = cols[3].strip()
            if not item_code or not yw or yw < min_yw:
                continue
            group_items.setdefault(group, set()).add(item_code)
            if min_week_yw is None or yw < min_week_yw:
                min_week_yw = yw

        if not group_items:
            return ['ดู Item ในกลุ่ม SKP', 'ดูแผนทอ Item', 'ดูเครื่องว่าง', 'ข้อมูลแผนทอ']

        try:
            min_w = int(min_week_yw[4:]) if min_week_yw and len(min_week_yw) == 6 else 0
        except ValueError:
            min_w = 0

        top_groups = sorted(group_items, key=lambda g: len(group_items[g]), reverse=True)[:2]
        suggestions = [f"Item ในกลุ่ม {g}" for g in top_groups]
        if min_w:
            suggestions.append(f"Item ทั้งหมดสัปดาห์ที่ {min_w}")
        suggestions.append("น้ำหนัก KP Weight ของ Item")
        return suggestions[:4]

    def _build_knit_plan_overview(self) -> str:
        item_cached, item_ready = get_data_cache().get('query_item')
        if not item_ready or not item_cached:
            return "ข้อมูลแผนทอยังไม่พร้อม กรุณาลองใหม่อีกครั้งค่ะ"
        item_data = item_cached.get('data', '')
        if not item_data:
            return "ไม่มีข้อมูลแผนทอในระบบค่ะ"

        min_yw = _min_plannable_yw()
        rows = []
        for line in item_data.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 4:
                continue
            item_code = cols[0].strip()
            group = cols[1].strip()
            yw = cols[3].strip()
            if not item_code or not yw or yw < min_yw:
                continue
            try:
                kp = float(cols[2].strip()) if cols[2].strip() else 0.0
            except ValueError:
                kp = 0.0
            rows.append({'item': item_code, 'group': group, 'kp': kp, 'yw': yw})

        if not rows:
            return "ไม่พบข้อมูลแผนทอในช่วงสัปดาห์ที่วางแผนได้ค่ะ"

        total_records = len(rows)
        distinct_items = set(r['item'] for r in rows)
        distinct_groups = set(r['group'] for r in rows)
        distinct_weeks = set(r['yw'] for r in rows)
        total_kp = sum(r['kp'] for r in rows)

        min_week_in_data = min(distinct_weeks)
        max_week_in_data = max(distinct_weeks)
        try:
            min_w = int(min_week_in_data[4:]) if len(min_week_in_data) == 6 else int(min_week_in_data)
            max_w = int(max_week_in_data[4:]) if len(max_week_in_data) == 6 else int(max_week_in_data)
        except ValueError:
            min_w, max_w = 0, 0

        # Per-group aggregation
        group_items: dict = {}
        group_kp: dict = {}
        for r in rows:
            g = r['group']
            group_items.setdefault(g, set()).add(r['item'])
            group_kp[g] = group_kp.get(g, 0.0) + r['kp']

        # Per-week aggregation
        week_kp: dict = {}
        for r in rows:
            week_kp[r['yw']] = week_kp.get(r['yw'], 0.0) + r['kp']

        top_group = max(group_kp, key=lambda g: group_kp[g]) if group_kp else '-'
        busiest_yw = max(week_kp, key=lambda w: week_kp[w]) if week_kp else '-'
        try:
            busiest_week_num = int(busiest_yw[4:]) if len(busiest_yw) == 6 else int(busiest_yw)
        except ValueError:
            busiest_week_num = 0

        # Sort groups by KP_Weight descending
        sorted_groups = sorted(group_kp.keys(), key=lambda g: group_kp[g], reverse=True)

        out = [
            f"ตอนนี้ระบบมีข้อมูลแผนทอทั้งหมด **{total_records:,} รายการ**",
            f"ครอบคลุมตั้งแต่สัปดาห์ที่ **{min_w}** ถึง **{max_w}** ({len(distinct_weeks)} สัปดาห์)",
            "",
            "**ภาพรวมแผนทอ**",
            f"- จำนวน Item ในแผนทั้งหมด : **{len(distinct_items)} Item**",
            f"- น้ำหนักแผนทอรวม : **{total_kp:,.0f} kg**",
            f"- จำนวนกลุ่มเครื่องที่เกี่ยวข้อง : **{len(distinct_groups)} กลุ่ม**",
            "",
            "**สรุปตามกลุ่มเครื่องหลัก**",
        ]

        for g in sorted_groups[:6]:
            n_items = len(group_items[g])
            kp = int(group_kp[g])
            label = g if g else '(ไม่ระบุกลุ่ม)'
            out.append(f"- **{label}** : {n_items} Item / {kp:,.0f} kg")

        out += [
            "",
            "จากภาพรวมตอนนี้",
            f"กลุ่มที่มีแผนทอมากที่สุดคือ **{top_group}** ({len(group_items.get(top_group, set()))} Item / {group_kp.get(top_group, 0):,.0f} kg)",
            f"สัปดาห์ที่มีแผนทอหนักที่สุดคือ สัปดาห์ที่ **{busiest_week_num}** ({week_kp.get(busiest_yw, 0):,.0f} kg)",
            "ถ้าต้องการดูรายละเอียดเพิ่มเติม สามารถระบุ Item, กลุ่มเครื่อง หรือสัปดาห์ที่ต้องการได้เลยค่ะ",
        ]
        return '\n'.join(out)

    def _get_knit_plan_suggestions(self) -> list:
        item_cached, item_ready = get_data_cache().get('query_item')
        if not item_ready or not item_cached:
            return ['แผนทอสัปดาห์นี้', 'ดูแผนทอกลุ่ม SKP', 'ดูเครื่องว่าง', 'ข้อมูล Item']

        item_data = item_cached.get('data', '')
        min_yw = _min_plannable_yw()

        group_kp: dict = {}
        week_kp: dict = {}
        for line in item_data.splitlines()[1:]:
            cols = line.split(',')
            if len(cols) < 4:
                continue
            yw = cols[3].strip()
            g = cols[1].strip()
            if not yw or yw < min_yw:
                continue
            try:
                kp = float(cols[2].strip()) if cols[2].strip() else 0.0
            except ValueError:
                kp = 0.0
            group_kp[g] = group_kp.get(g, 0.0) + kp
            week_kp[yw] = week_kp.get(yw, 0.0) + kp

        top_groups = sorted(group_kp, key=lambda g: group_kp[g], reverse=True)[:2]
        busiest_yw = max(week_kp, key=lambda w: week_kp[w]) if week_kp else None

        suggestions = [f"แผนทอกลุ่ม {g}" for g in top_groups]
        if busiest_yw:
            try:
                wn = int(busiest_yw[4:]) if len(busiest_yw) == 6 else int(busiest_yw)
                suggestions.append(f"แผนทอสัปดาห์ที่ {wn}")
            except ValueError:
                pass
        suggestions.append("ดูเครื่องว่างในแผนทอ")
        return suggestions[:4]

    async def process_message(
        self,
        user_message: str,
        username: str = 'คุณ',  # noqa: ARG002 — kept for API compatibility
        conversation_history: Optional[List[Dict]] = None,
        image_base64: Optional[str] = None,
        image_media_type: str = 'image/png',
    ) -> ProcessedResponse:
        if user_message.strip() == 'ข้อมูลเครื่องจักร':
            return ProcessedResponse(
                message=self._build_machine_capacity_overview(),
                response_type='text',
                processing_path='shortcut',
                metadata={'intent': 'machine_info_menu', 'confidence': 1.0, 'matched_keywords': []},
                suggestions=self._get_machine_capacity_suggestions(),
            )

        if user_message.strip() == 'ข้อมูล Capacity':
            return ProcessedResponse(
                message=self._build_capacity_overview(),
                response_type='text',
                processing_path='shortcut',
                metadata={'intent': 'capacity_menu', 'confidence': 1.0, 'matched_keywords': []},
                suggestions=self._get_capacity_suggestions(),
            )

        if user_message.strip() == 'ข้อมูลแผนทอ':
            return ProcessedResponse(
                message=self._build_knit_plan_overview(),
                response_type='text',
                processing_path='shortcut',
                metadata={'intent': 'knitting_plan', 'confidence': 1.0, 'matched_keywords': []},
                suggestions=self._get_knit_plan_suggestions(),
            )

        if user_message.strip() == 'ข้อมูล Item':
            return ProcessedResponse(
                message=self._build_item_overview(),
                response_type='text',
                processing_path='shortcut',
                metadata={'intent': 'item_data', 'confidence': 1.0, 'matched_keywords': []},
                suggestions=self._get_item_suggestions(),
            )

        history_msgs = self._build_history_messages(conversation_history or [])
        _now = datetime.now()
        system_prompt = SYSTEM_PROMPT_TEMPLATE.format(
            min_yw=_min_plannable_yw(),
            today=_now.strftime('%Y-%m-%d'),
            current_yw=_yw_from_date(_now),
        )
        lang = _detect_reply_language(user_message, conversation_history)
        lang_instruction = "IMPORTANT: Reply in Thai language." if lang == 'thai' else "IMPORTANT: Reply in English."

        if image_base64:
            user_msg_content = [
                {"type": "image_url", "image_url": {"url": f"data:{image_media_type};base64,{image_base64}"}},
                {"type": "text", "text": user_message or "กรุณาวิเคราะห์รูปภาพนี้"},
            ]
            image_system = [{"role": "system", "content": "User has shared a screenshot or image. Analyze it and answer based on what you see. If it shows I-SAVE related data (tables, schedules, bookings), respond accordingly."}]
        else:
            user_msg_content = user_message
            image_system = []

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "system", "content": lang_instruction},
            *image_system,
            *_export_nudge(user_message),
            *history_msgs,
            {"role": "user", "content": user_msg_content},
        ]

        tool_calls_log: List[Dict] = []
        chart_spec: Optional[Dict] = None
        table_spec: Optional[Dict] = None
        total_prompt_tokens = 0
        total_completion_tokens = 0
        loop = asyncio.get_running_loop()

        for _ in range(MAX_TOOL_ITERATIONS):
            response = None
            last_error: Optional[Exception] = None
            for attempt in range(OPENAI_MAX_RETRIES + 1):
                try:
                    response = await loop.run_in_executor(
                        None,
                        lambda: self._openai.chat.completions.create(
                            model=self._model,
                            messages=messages,
                            tools=TOOLS,
                            tool_choice="auto",
                            max_completion_tokens=4096,
                        ),
                    )
                    break
                except Exception as e:
                    last_error = e
                    logger.error(f"OpenAI API error (attempt {attempt + 1}/{OPENAI_MAX_RETRIES + 1}): {e}")
                    if attempt < OPENAI_MAX_RETRIES:
                        await asyncio.sleep(OPENAI_RETRY_BACKOFF * (2 ** attempt))

            if response is None:
                # ลองครบทุกครั้งแล้วยังพลาด — ซ่อนรายละเอียด error ดิบจาก user (log ไว้แล้วด้านบน)
                logger.error(f"OpenAI API failed after {OPENAI_MAX_RETRIES + 1} attempts: {last_error}")
                err_msg = (
                    "Sorry, the AI system is temporarily unavailable. Please try again shortly."
                    if lang == 'english'
                    else "ขออภัยค่ะ ระบบ AI ขัดข้องชั่วคราว กรุณาลองใหม่อีกครั้งในอีกสักครู่นะคะ"
                )
                return ProcessedResponse(
                    message=err_msg,
                    response_type='text',
                    metadata={'intent': 'error', 'confidence': 0.0, 'matched_keywords': []},
                )

            choice = response.choices[0]
            msg = choice.message
            if response.usage:
                total_prompt_tokens += response.usage.prompt_tokens
                total_completion_tokens += response.usage.completion_tokens

            if choice.finish_reason == "tool_calls" and msg.tool_calls:
                messages.append({
                    "role": "assistant",
                    "content": msg.content or "",
                    "tool_calls": [
                        {
                            "id": tc.id,
                            "type": "function",
                            "function": {"name": tc.function.name, "arguments": tc.function.arguments},
                        }
                        for tc in msg.tool_calls
                    ],
                })
                for tool_call in msg.tool_calls:
                    if tool_call.function.name == "render_chart":
                        spec, result = self._build_chart_spec(tool_call.function.arguments)
                        if spec:
                            chart_spec = spec
                    elif tool_call.function.name == "render_table":
                        spec, result = self._build_table_spec(tool_call.function.arguments)
                        if spec:
                            table_spec = spec
                    else:
                        result = self._execute_tool_call(tool_call)
                    tool_calls_log.append({
                        'tool': tool_call.function.name,
                        'args': tool_call.function.arguments,
                        'result_rows': len(result.splitlines()),
                    })
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tool_call.id,
                        "content": result,
                    })
            else:
                fallback = "Sorry, I'm unable to respond." if lang == 'english' else "ขออภัยค่ะ ไม่สามารถตอบได้"
                final_message = _clean_response(msg.content or fallback)
                alert = self._build_capacity_alert(tool_calls_log, lang)
                if alert:
                    final_message = final_message + alert
                _data = {}
                if chart_spec:
                    _data['chart'] = chart_spec
                if table_spec:
                    _data['table'] = table_spec
                return ProcessedResponse(
                    message=final_message,
                    response_type='chart' if chart_spec else ('table' if table_spec else 'text'),
                    data=_data or None,
                    processing_path='agent',
                    mcp_calls=tool_calls_log,
                    metadata={
                        'intent': 'agent',
                        'confidence': 1.0,
                        'matched_keywords': [c['tool'] for c in tool_calls_log],
                        'model': self._model,
                        'prompt_tokens': total_prompt_tokens,
                        'completion_tokens': total_completion_tokens,
                        'total_tokens': total_prompt_tokens + total_completion_tokens,
                    },
                    suggestions=self._get_suggestions(tool_calls_log, lang),
                )

        timeout_msg = "Sorry, the system took too long to respond. Please try again." if lang == 'english' else "ขออภัยค่ะ ระบบประมวลผลนานเกินไป กรุณาถามใหม่อีกครั้ง"
        return ProcessedResponse(
            message=timeout_msg,
            response_type='text',
            processing_path='agent',
            mcp_calls=tool_calls_log,
            metadata={'intent': 'agent', 'confidence': 0.0, 'matched_keywords': []},
        )

    async def process_message_stream(
        self,
        user_message: str,
        username: str = 'คุณ',
        conversation_history: Optional[List[Dict]] = None,
        image_base64: Optional[str] = None,
        image_media_type: str = 'image/png',
    ):
        """Async generator yielding SSE event dicts for streaming response."""

        # Shortcuts — yield single done event immediately (no streaming needed)
        _shortcuts = {
            'ข้อมูลเครื่องจักร': (self._build_machine_capacity_overview, self._get_machine_capacity_suggestions, 'machine_info_menu'),
            'ข้อมูล Capacity':   (self._build_capacity_overview,         self._get_capacity_suggestions,         'capacity_menu'),
            'ข้อมูลแผนทอ':      (self._build_knit_plan_overview,         self._get_knit_plan_suggestions,        'knitting_plan'),
            'ข้อมูล Item':       (self._build_item_overview,              self._get_item_suggestions,             'item_data'),
        }
        shortcut = _shortcuts.get(user_message.strip())
        if shortcut:
            build_fn, sugg_fn, intent = shortcut
            yield {
                'type': 'done',
                'message': build_fn(),
                'suggestions': sugg_fn(),
                'chart': None,
                'processing_path': 'shortcut',
                'metadata': {'intent': intent, 'confidence': 1.0, 'matched_keywords': []},
                'mcp_calls': [],
            }
            return

        history_msgs = self._build_history_messages(conversation_history or [])
        _now = datetime.now()
        system_prompt = SYSTEM_PROMPT_TEMPLATE.format(
            min_yw=_min_plannable_yw(),
            today=_now.strftime('%Y-%m-%d'),
            current_yw=_yw_from_date(_now),
        )
        lang = _detect_reply_language(user_message, conversation_history)
        lang_instruction = "IMPORTANT: Reply in Thai language." if lang == 'thai' else "IMPORTANT: Reply in English."

        if image_base64:
            user_msg_content = [
                {"type": "image_url", "image_url": {"url": f"data:{image_media_type};base64,{image_base64}"}},
                {"type": "text", "text": user_message or "กรุณาวิเคราะห์รูปภาพนี้"},
            ]
            image_system = [{"role": "system", "content": "User has shared a screenshot or image. Analyze it and answer based on what you see. If it shows I-SAVE related data (tables, schedules, bookings), respond accordingly."}]
        else:
            user_msg_content = user_message
            image_system = []

        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "system", "content": lang_instruction},
            *image_system,
            *_export_nudge(user_message),
            *history_msgs,
            {"role": "user", "content": user_msg_content},
        ]

        tool_calls_log: List[Dict] = []
        chart_spec: Optional[Dict] = None
        table_spec: Optional[Dict] = None
        total_prompt_tokens = 0
        total_completion_tokens = 0

        for _ in range(MAX_TOOL_ITERATIONS):
            stream = None
            last_error = None
            for attempt in range(OPENAI_MAX_RETRIES + 1):
                try:
                    stream = await self._openai_async.chat.completions.create(
                        model=self._model,
                        messages=messages,
                        tools=TOOLS,
                        tool_choice="auto",
                        max_completion_tokens=4096,
                        stream=True,
                    )
                    break
                except Exception as e:
                    last_error = e
                    logger.error(f"OpenAI stream error (attempt {attempt + 1}): {e}")
                    if attempt < OPENAI_MAX_RETRIES:
                        await asyncio.sleep(OPENAI_RETRY_BACKOFF * (2 ** attempt))

            if stream is None:
                err_msg = "Sorry, the AI system is temporarily unavailable." if lang == 'english' else "ขออภัยค่ะ ระบบ AI ขัดข้องชั่วคราว กรุณาลองใหม่อีกครั้งนะคะ"
                yield {'type': 'error', 'message': err_msg}
                return

            content_parts: List[str] = []
            tool_call_accum: Dict[int, Dict] = {}
            finish_reason = None

            async for chunk in stream:
                if not chunk.choices:
                    if hasattr(chunk, 'usage') and chunk.usage:
                        total_prompt_tokens += chunk.usage.prompt_tokens or 0
                        total_completion_tokens += chunk.usage.completion_tokens or 0
                    continue
                choice = chunk.choices[0]
                finish_reason = choice.finish_reason or finish_reason
                delta = choice.delta

                if delta.content:
                    content_parts.append(delta.content)
                    yield {'type': 'token', 'text': delta.content}

                if delta.tool_calls:
                    for tc in delta.tool_calls:
                        idx = tc.index
                        if idx not in tool_call_accum:
                            tool_call_accum[idx] = {'id': '', 'name': '', 'arguments': ''}
                        if tc.id:
                            tool_call_accum[idx]['id'] = tc.id
                        if tc.function:
                            if tc.function.name:
                                tool_call_accum[idx]['name'] += tc.function.name
                            if tc.function.arguments:
                                tool_call_accum[idx]['arguments'] += tc.function.arguments

                if hasattr(chunk, 'usage') and chunk.usage:
                    total_prompt_tokens += chunk.usage.prompt_tokens or 0
                    total_completion_tokens += chunk.usage.completion_tokens or 0

            if tool_call_accum:  # handle both 'tool_calls' and 'length' (truncated JSON)
                messages.append({
                    "role": "assistant",
                    "content": ''.join(content_parts),
                    "tool_calls": [
                        {"id": tc['id'], "type": "function",
                         "function": {"name": tc['name'], "arguments": tc['arguments']}}
                        for tc in tool_call_accum.values()
                    ],
                })
                for idx in sorted(tool_call_accum.keys()):
                    tc = tool_call_accum[idx]
                    tool_name = tc['name']
                    yield {'type': 'status', 'text': _TOOL_STATUS_LABELS.get(tool_name, 'กำลังประมวลผล...')}

                    if tool_name == 'render_chart':
                        spec, result = self._build_chart_spec(tc['arguments'])
                        if spec:
                            chart_spec = spec
                    elif tool_name == 'render_table':
                        spec, result = self._build_table_spec(tc['arguments'])
                        if spec:
                            table_spec = spec
                    else:
                        _fake = type('TC', (), {
                            'function': type('F', (), {'name': tool_name, 'arguments': tc['arguments']})()
                        })()
                        result = self._execute_tool_call(_fake)
                        tool_calls_log.append({'tool': tool_name, 'args': tc['arguments'], 'result_rows': len(result.splitlines())})

                    messages.append({"role": "tool", "tool_call_id": tc['id'], "content": result})
                continue

            # finish_reason == 'stop' — all tokens already streamed
            full_message = _clean_response(''.join(content_parts))
            alert = self._build_capacity_alert(tool_calls_log, lang)
            if alert:
                full_message = full_message + alert
                yield {'type': 'token', 'text': alert}
            yield {
                'type': 'done',
                'message': full_message,
                'chart': chart_spec,
                'table': table_spec,
                'processing_path': 'agent',
                'metadata': {
                    'intent': 'agent',
                    'confidence': 1.0,
                    'matched_keywords': [c['tool'] for c in tool_calls_log],
                    'model': self._model,
                    'prompt_tokens': total_prompt_tokens,
                    'completion_tokens': total_completion_tokens,
                    'total_tokens': total_prompt_tokens + total_completion_tokens,
                },
                'mcp_calls': tool_calls_log,
            }
            return

        timeout_msg = "Sorry, the system took too long. Please try again." if lang == 'english' else "ขออภัยค่ะ ระบบประมวลผลนานเกินไป กรุณาถามใหม่อีกครั้ง"
        yield {
            'type': 'done', 'message': timeout_msg, 'chart': None,
            'processing_path': 'agent', 'mcp_calls': [],
            'metadata': {'intent': 'timeout', 'confidence': 0.0, 'matched_keywords': []},
        }

    def _build_capacity_alert(self, tool_calls_log: list, lang: str = 'thai') -> str:
        """Return warning text if queried group(s) > 85% utilization."""
        tools_used = {c['tool'] for c in (tool_calls_log or [])}
        if not (tools_used & {'get_machine_capacity', 'get_booking'}):
            return ''

        # Collect group filters used in the query — only alert for those groups
        queried_groups: set = set()
        for call in tool_calls_log:
            if call['tool'] in ('get_machine_capacity', 'get_booking'):
                try:
                    args = json.loads(call['args']) if isinstance(call['args'], str) else {}
                    g = (args.get('group') or '').strip().lower()
                    if g:
                        queried_groups.add(g)
                except Exception:
                    pass

        parsed = self._parse_mc_snapshot()
        if not parsed:
            return ''
        _, _, week_num, group_total, group_used, _ = parsed

        critical = {}
        for g in group_total:
            if group_total.get(g, 0) <= 0:
                continue
            # If query had a group filter, only check that group
            if queried_groups and not _group_matches(g.lower(), list(queried_groups)):
                continue
            pct = group_used[g] / group_total[g] * 100
            if pct > 85.0:
                critical[g] = pct

        if not critical:
            return ''
        ranked = sorted(critical.items(), key=lambda x: x[1], reverse=True)
        if lang == 'english':
            lines = [f"\n⚠️ High utilization alert — week {week_num}:"]
            for g, pct in ranked:
                lines.append(f"• **{g}** — {round(pct, 1)}% utilized")
            lines.append("Consider planning ahead or checking available slots.")
        else:
            lines = [f"\n⚠️ แจ้งเตือน Capacity ตึง สัปดาห์ที่ {week_num}:"]
            for g, pct in ranked:
                lines.append(f"• **{g}** — ใช้งาน **{round(pct, 1)}%**")
            lines.append("แนะนำให้วางแผนล่วงหน้าหรือตรวจสอบเครื่องว่างค่ะ")
        return '\n'.join(lines)

    def _get_suggestions(self, tool_calls_log: list, lang: str = 'thai') -> list:
        tools_used = {c['tool'] for c in tool_calls_log}
        if lang == 'english':
            if 'get_item_plan' in tools_used or 'get_knit_plan' in tools_used:
                return ['Check available machines for this item', 'View all KP Weights', 'View knit plan this week']
            if 'get_machine_capacity' in tools_used:
                return ['Check available machines next week', 'View booking', 'View knit plan']
            if 'get_booking' in tools_used:
                return ['Check available machines', 'View knit plan', 'View item plan']
            return ['Machine info', 'Item info', 'Capacity info']
        if 'get_item_plan' in tools_used or 'get_knit_plan' in tools_used:
            return ['ดูเครื่องว่างสำหรับ item นี้', 'ดู KP Weight ทั้งหมด', 'ดูแผนทอสัปดาห์นี้']
        if 'get_machine_capacity' in tools_used:
            return ['ดูเครื่องว่างสัปดาห์ถัดไป', 'ดูการจอง', 'ดูแผนทอ']
        if 'get_booking' in tools_used:
            return ['ดูเครื่องว่าง', 'ดูแผนทอ', 'ดู item plan']
        return ['ข้อมูลเครื่องจักร', 'ข้อมูล Item', 'ข้อมูล Capacity']


_response_processor_instance = None


def get_response_processor() -> ResponseProcessor:
    global _response_processor_instance
    if _response_processor_instance is None:
        _response_processor_instance = ResponseProcessor()
    return _response_processor_instance
